import os
import pandas as pd
import openpyxl
from core.parser_dpgf import parse_dpgf
from core.parser_tco import parse_tco
from core.merger import merge_company_into_tco
import tempfile
from pathlib import Path

tmp_path = Path(tempfile.mkdtemp())

# 1. Create TCO
tco_path = tmp_path / "valid_tco.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["PROJET TEST"])
ws.append(["Code", "Désignation", "Qu.", "U", "Px U", "Px Tot"])
ws.append(["01", "Section 1", "", "", "", ""])
ws.append(["01.1", "Article 1", "10", "u", "10", "100"])
wb.save(tco_path)

# 2. Create DPGF
dpgf_path = tmp_path / "valid_dpgf.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Code", "Désignation", "Qu.", "U", "Px U", "Px Tot"])
ws.append(["01.1", "Article 1 - DPGF", "10", "u", "15", "150"])
wb.save(dpgf_path)

# 3. Create Empty DPGF
empty_path = tmp_path / "empty_dpgf.xlsx"
wb = openpyxl.Workbook()
wb.save(empty_path)

print("\n--- RUNNING test_merge_valid ---")
try:
    tco_df, _ = parse_tco(str(tco_path))
    dpgf_df, _ = parse_dpgf(str(dpgf_path))
    merged_df, alerts = merge_company_into_tco(tco_df, dpgf_df, "TEST_COMP")
    
    article_row = merged_df[merged_df["Code"] == "01.1"].iloc[0]
    print(f"Article 01.1 Px_Tot_HT = {article_row.get('TEST_COMP_Px_Tot_HT')}")
    assert article_row["TEST_COMP_Px_Tot_HT"] == 150, f"Expected 150, got {article_row.get('TEST_COMP_Px_Tot_HT')}"
    
    section_row = merged_df[merged_df["Code"] == "01"].iloc[0]
    print(f"Section 01 Px_Tot_HT = {section_row.get('TEST_COMP_Px_Tot_HT')}")
    assert section_row["TEST_COMP_Px_Tot_HT"] == 150, f"Expected 150, got {section_row.get('TEST_COMP_Px_Tot_HT')}"
    print("SUCCESS test_merge_valid")
except Exception as e:
    import traceback
    traceback.print_exc()

print("\n--- RUNNING test_merge_empty_dpgf ---")
try:
    dpgf_empty_df, _ = parse_dpgf(str(empty_path))
    merged_empty_df, empty_alerts = merge_company_into_tco(tco_df, dpgf_empty_df, "ENTREPRISE_X")
    
    assert not merged_empty_df.empty, "Merged DF is empty!"
    assert "ENTREPRISE_X_Px_Tot_HT" in merged_empty_df.columns, "Company columns missing!"
    print(f"Alerts: {empty_alerts}")
    assert len(empty_alerts) == 1
    print("SUCCESS test_merge_empty_dpgf")
except Exception as e:
    import traceback
    traceback.print_exc()
