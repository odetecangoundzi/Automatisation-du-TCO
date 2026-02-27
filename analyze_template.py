import openpyxl

XLSX_PATH = (
    r"s:\TRANSFORMATION DIGITALE\Automatisation du TCO"
    r"\TCO_APP\Template_DPGF\TCO_MODELE"
    r"\TCO 01 - TERRASSEMENT  VRD  PAYSAGE.xlsx"
)

wb = openpyxl.load_workbook(XLSX_PATH)

print("=== NAMED STYLES ===")
for ns in wb.named_styles:
    print(f"  NamedStyle: {ns}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"=== FEUILLE: {sheet_name} ===")
    print(f"{'='*80}")
    print(f"Dimensions: {ws.dimensions}")
    print(
        f"Min row: {ws.min_row}, Max row: {ws.max_row}, "
        f"Min col: {ws.min_column}, Max col: {ws.max_column}"
    )
    print(f"Freeze panes: {ws.freeze_panes}")

    print("\n--- Merged cells ---")
    for mc in ws.merged_cells.ranges:
        print(f"  {mc}")

    print("\n--- Largeurs de colonnes ---")
    for col_letter, cd in ws.column_dimensions.items():
        print(f"  Col {col_letter}: width={cd.width}, hidden={cd.hidden}")

    print("\n--- Hauteurs de lignes ---")
    for row_idx, rd in ws.row_dimensions.items():
        if rd.height:
            print(f"  Row {row_idx}: height={rd.height}, hidden={rd.hidden}")

    print("\n--- Cellules (30 premieres lignes) ---")
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            has_value = cell.value is not None

            fill = cell.fill
            try:
                fg_color = fill.fgColor
                if fg_color.type == "rgb":
                    fill_info = fg_color.rgb
                elif fg_color.type == "theme":
                    fill_info = f"theme:{fg_color.theme},tint:{fg_color.tint}"
                else:
                    fill_info = fg_color.type
            except Exception as exc:
                fill_info = f"err:{exc}"

            font = cell.font
            try:
                if font.color.type == "rgb":
                    font_color = font.color.rgb
                elif font.color.type == "theme":
                    font_color = (
                        f"theme:{font.color.theme},tint:{font.color.tint}"
                    )
                else:
                    font_color = font.color.type
            except Exception as exc:
                font_color = f"err:{exc}"

            border = cell.border
            border_parts = []
            for side_name in ["left", "right", "top", "bottom"]:
                side = getattr(border, side_name)
                if side and side.border_style and side.border_style != "none":
                    try:
                        if side.color.type == "rgb":
                            bc = side.color.rgb
                        else:
                            bc = f"theme:{side.color.theme}"
                    except Exception:
                        bc = "?"
                    border_parts.append(
                        f"{side_name}:{side.border_style}({bc})"
                    )
            border_str = "|".join(border_parts) if border_parts else "none"

            merge_type = ""
            for mc_range in ws.merged_cells.ranges:
                in_range = (
                    mc_range.min_row <= cell.row <= mc_range.max_row
                    and mc_range.min_col <= cell.column <= mc_range.max_col
                )
                if in_range:
                    if (
                        cell.row == mc_range.min_row
                        and cell.column == mc_range.min_col
                    ):
                        merge_type = f"MERGE_MASTER({mc_range})"
                    else:
                        merge_type = f"MERGE_SLAVE({mc_range})"
                    break

            val_repr = repr(cell.value)
            if len(val_repr) > 60:
                val_repr = val_repr[:57] + "..."

            skip_fill = fill_info in ("00000000", "none", "theme:0,tint:0.0")
            if has_value or not skip_fill:
                print(
                    f"  {cell.coordinate:6s} | val={val_repr:62s} | "
                    f"fill={fill_info:30s} | "
                    f"bold={str(font.bold):5s} | size={str(font.size):5s} | "
                    f"color={font_color:25s} | "
                    f"fmt={cell.number_format!r:20s} | "
                    f"align={str(cell.alignment.horizontal):10s} | "
                    f"border={border_str:50s} | {merge_type}"
                )

print("\n\nDone.")
