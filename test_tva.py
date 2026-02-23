import sys, os
import io
import openpyxl
from core.parser_tco import parse_tco
from core.parser_dpgf import parse_dpgf
from core.merger import merge_company_into_tco
from core.exporter import export_tco

def test_tva_rate(rate):
    print(f"Testing TVA rate: {rate}")
    tco_df, meta = parse_tco(r"d:\CTO\TCO_FINAL.xlsx")
    dpgf_df, _ = parse_dpgf(r"d:\CTO\DPGF LOT 01 - DESAMIANTAGE - CURAGE - GROS OEUVRE.xlsx")
    
    merged, _ = merge_company_into_tco(tco_df, dpgf_df, "TEST_COMP", tva_rate=rate)
    buffer = export_tco(merged, meta, output_path=None, alerts=[], tva_rate=rate)
    
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    
    found_tva = False
    for r in range(1, ws.max_row + 1):
        desig = str(ws.cell(r, 2).value).lower()
        if "tva" in desig and "ht" not in desig:
            formula_f = str(ws.cell(r, 6).value)
            formula_i = str(ws.cell(r, 9).value)
            print(f"Row {r} ({desig}): Col F={formula_f}, Col I={formula_i}")
            
            if str(rate) in formula_f and str(rate) in formula_i:
                found_tva = True
                print(f"[OK] Found rate {rate} in row {r}")
                break
            
    if not found_tva:
        print("DEBUG: All 'total' or 'tva' rows found:")
        for r in range(1, ws.max_row + 1):
            d = str(ws.cell(r, 2).value).lower()
            if "total" in d or "tva" in d:
                print(f"Row {r}: '{d}' | F: {ws.cell(r, 6).value} | I: {ws.cell(r, 9).value}")

    assert found_tva, f"TVA rate {rate} not found in any formula"
    print(f"[OK] TVA rate {rate} verified")

if __name__ == "__main__":
    try:
        test_tva_rate(0.055)
        test_tva_rate(0.10)
        test_tva_rate(0.20)
        print("\nALL TVA TESTS PASSED")
    except Exception as e:
        print(f"\n[FAIL] TEST FAILED: {str(e)}")
        # sys.exit(1) # Don't exit to allow seeing the output
