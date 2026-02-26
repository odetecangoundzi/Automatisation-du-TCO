"""Test script for Export du TCO end-to-end."""
import sys
import os

sys.path.insert(0, r"d:\CTO")

from core.parser_tco import parse_tco
from core.parser_dpgf import parse_dpgf
from core.merger import merge_company_into_tco
from core.exporter import export_tco

print("=== Step 1: Parse TCO ===")
tco_df, meta = parse_tco(r"d:\CTO\DPGF LOT 01 - DESAMIANTAGE - CURAGE - GROS OEUVRE.xlsx")
data_count = len(tco_df[tco_df["row_type"] == "data"])
print(f"TCO: {len(tco_df)} rows, {data_count} data rows")
print(f"Project: {meta['project_info']}")
print(f"Header row: {meta['header_row']}")
print()

print("=== Step 2: Parse DPGF ===")
dpgf_df, alerts = parse_dpgf(r"d:\CTO\MAB_SUD_OUEST.xlsx")
dpgf_data_count = len(dpgf_df[dpgf_df["row_type"] == "data"])
print(f"DPGF: {len(dpgf_df)} rows, {dpgf_data_count} data rows")
print(f"Alerts: {len(alerts)}")
for a in alerts[:8]:
    print(f"  {a['type']}: {a['code']} - {a['message']}")
if len(alerts) > 8:
    print(f"  ... and {len(alerts)-8} more")
print()

print("=== Step 3: Merge ===")
merged, merge_alerts = merge_company_into_tco(tco_df, dpgf_df, "MAB SUD-OUEST")
print(f"Merged: {len(merged)} rows, {len(merged.columns)} columns")
print(f"Merge alerts: {len(merge_alerts)}")
for a in merge_alerts[:5]:
    print(f"  {a['type']}: {a['code']} - {a['message']}")
print()

print("=== Step 4: Export ===")
output = r"d:\CTO\outputs\test_output.xlsx"
export_tco(merged, meta, output, alerts + merge_alerts)
print(f"Exported to: {output}")
print()

# Verify output
import openpyxl
wb = openpyxl.load_workbook(output, data_only=True)
ws = wb.active
print(f"=== Output Verification ===")
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print(f"Row 1 (headers): ", end="")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        from openpyxl.utils import get_column_letter
        print(f"{get_column_letter(col)}={val}", end=" | ")
print()
print(f"Row 2 (col names): ", end="")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=2, column=col).value
    if val:
        from openpyxl.utils import get_column_letter
        print(f"{get_column_letter(col)}={val}", end=" | ")
print()

# Check MAB SUD-OUEST data for row 01.2.1.1.1
for row_idx in range(3, ws.max_row + 1):
    code = ws.cell(row=row_idx, column=1).value
    if code and str(code).strip() == "01.2.1.1.1":
        print(f"\nRow {row_idx} (01.2.1.1.1):")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                print(f"  {get_column_letter(col)}: {val}")
        break

wb.close()
print("\n=== ALL TESTS PASSED ===")
