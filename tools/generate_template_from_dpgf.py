"""
Générateur de template TCO depuis un DPGF entreprise.

Usage:
    python tools/generate_template_from_dpgf.py <dpgf.xlsx> [options]

Options:
    --output        Chemin de sortie (.xlsx)
    --lot           Identifiant du lot (ex: Lot12) — auto-détecté si absent
    --name          Nom descriptif (ex: EQUIPEMENTS CUISINE)
    --keep-prices   Conserver les prix du DPGF comme référence
    --no-validate   Ne pas valider le template après génération

Exemples:
    python tools/generate_template_from_dpgf.py \\
        "Template_DPGF/LOT 12/LOT 12_EQUIP_CUISINE_DPGF.xlsx"
    python tools/generate_template_from_dpgf.py mon_dpgf.xlsx \\
        --lot Lot07 --name "VRD TERRASSEMENT"
    python tools/generate_template_from_dpgf.py mon_dpgf.xlsx --keep-prices

Le fichier généré peut être chargé dans l'app TCO comme "TCO Modèle".
"""
from __future__ import annotations

import argparse
import os
import sys

# sys.path doit être modifié avant les imports locaux (E402 inévitable ici)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402

from core.parser_dpgf import parse_dpgf  # noqa: E402
from logger import get_logger  # noqa: E402

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Styles (cohérents avec exporter.py)
# ---------------------------------------------------------------------------

FONT_HEADER = Font(name="Tahoma", bold=True, size=10, color="FFFFFF")
FONT_SECTION = Font(name="Tahoma", bold=True, size=11, color="AC2C18")
FONT_RECAP = Font(name="Tahoma", bold=True, size=11, color="000000")
FONT_SUB = Font(name="Tahoma", bold=True, size=9, color="314E85")
FONT_DATA = Font(name="Tahoma", size=9, color="000000")
FONT_TOTAL = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")

FILL_HEADER = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
FILL_WHITE = PatternFill(
    start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
)
FILL_MONTANT_HT = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
)
FILL_TVA = PatternFill(
    start_color="2E75B6", end_color="2E75B6", fill_type="solid"
)
FILL_TTC = PatternFill(
    start_color="0D2137", end_color="0D2137", fill_type="solid"
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MONEY_FORMAT = r'###,###,###,##0.00\ \€;\-###,###,###,##0.00\ \€;'
QTY_FORMAT = r'###,###,###,##0.00;\-###,###,###,##0.00;'

# Colonne M (index 13, 1-based) = Entete — doit correspondre à idx_entete=12
# dans utils.py (0-based)
ENTETE_EXCEL_COL = 13


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _row_type_to_entete(row_type: str, lot: str) -> str:
    """Génère la valeur Entete correcte pour que classify_row() reconnaisse."""
    mapping = {
        "section_header": f"Bd_{lot}_Bord",
        "recap": f"Bord_{lot}_Recap",
        "recap_summary": f"RecapBord_{lot}",
        "sub_section": f"Ouv_{lot}_Niv1",
        "article": f"Ouv_{lot}_Art",
        "total_line": f"LignesTot_{lot}",
    }
    return mapping.get(row_type, "")


def _auto_detect_lot(df) -> str:
    """Détecte le numéro de lot depuis les codes du DPGF."""
    mask = df["Code"].ne("") & df["row_type"].isin(
        ["article", "section_header"]
    )
    non_empty = df[mask]
    if non_empty.empty:
        return "LotXX"
    first_code = str(non_empty.iloc[0]["Code"]).strip()
    if "." in first_code:
        return f"Lot{first_code.split('.')[0]}"
    return f"Lot{first_code[:4]}"


def _make_recap_row(section_code: str, lot: str) -> dict:
    """Crée une ligne recap pour une section."""
    return {
        "Code": "",
        "Désignation": f"Total {section_code}",
        "Qu.": None,
        "U": "",
        "Px_U_HT": None,
        "Px_Tot_HT": None,
        "row_type": "recap",
        "Entete": _row_type_to_entete("recap", lot),
        "parent_code": section_code,
    }


def _make_total_row(label: str, lot: str) -> dict:
    """Crée une ligne total (Montant HT, TVA, Montant TTC)."""
    return {
        "Code": "",
        "Désignation": label,
        "Qu.": None,
        "U": "",
        "Px_U_HT": None,
        "Px_Tot_HT": None,
        "row_type": "total_line",
        "Entete": _row_type_to_entete("total_line", lot),
        "parent_code": "",
    }


def _build_final_rows(df, lot: str) -> list[dict]:
    """
    Reconstruit les lignes ordonnées avec :
    - Gestion des codes de section dupliqués (détail + récap synthèse).
      Stratégie : on garde la PREMIERE occurrence (détail complet).
    - Lignes recap injectées si absentes.
    - Lignes total HT / TVA / TTC ajoutées en fin.
    """
    seen_section_codes: set[str] = set()
    sections: list[str] = []
    section_data: dict[str, dict] = {}
    current_section: str | None = None
    total_rows: list[dict] = []

    for _, row in df.iterrows():
        rt = row["row_type"]
        code = str(row.get("Code", "")).strip()

        if rt == "empty":
            continue

        row_dict = row.to_dict()

        if rt == "section_header":
            if code in seen_section_codes:
                # Doublon — deuxième bloc (récap synthèse) → on ignore
                current_section = None
                continue
            seen_section_codes.add(code)
            current_section = code
            sections.append(code)
            section_data[code] = {
                "header": row_dict,
                "content": [],
                "has_recap": False,
            }

        elif rt in ("total_line", "total_text"):
            total_rows.append(row_dict)

        elif current_section is not None:
            if rt == "recap":
                section_data[current_section]["has_recap"] = True
            if rt != "recap_summary":
                section_data[current_section]["content"].append(row_dict)

    # Construire la liste finale
    final_rows: list[dict] = []

    for sec_code in sections:
        sec = section_data[sec_code]
        final_rows.append(sec["header"])

        for r in sec["content"]:
            if r["row_type"] not in ("recap", "total_line", "total_text"):
                final_rows.append(r)

        if sec["has_recap"]:
            for r in sec["content"]:
                if r["row_type"] == "recap":
                    final_rows.append(r)
        else:
            final_rows.append(_make_recap_row(sec_code, lot))

    # Total lines : issues du DPGF ou générées
    has_ht = any(
        "montant ht" in str(r.get("Désignation", "")).lower()
        for r in total_rows
    )
    if total_rows and has_ht:
        final_rows.extend(total_rows)
    else:
        final_rows.append(_make_total_row("Montant HT", lot))
        final_rows.append(_make_total_row("TVA 20%", lot))
        final_rows.append(_make_total_row("Montant TTC", lot))

    return final_rows


def _style_row(ws, excel_row: int, row_type: str, max_col: int = 6) -> None:
    """Applique police et bordure à une ligne."""
    font_map = {
        "section_header": FONT_SECTION,
        "recap": FONT_RECAP,
        "recap_summary": FONT_RECAP,
        "sub_section": FONT_SUB,
        "total_line": FONT_TOTAL,
    }
    font = font_map.get(row_type, FONT_DATA)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=excel_row, column=c)
        cell.font = font
        cell.fill = FILL_WHITE
        cell.border = THIN_BORDER


def _apply_total_fill(ws, excel_row: int, desig: str, max_col: int = 6) -> None:
    """Applique le fond coloré sur les lignes total."""
    dl = desig.lower()
    if "montant ht" in dl:
        fill = FILL_MONTANT_HT
    elif "tva" in dl and "ht" not in dl:
        fill = FILL_TVA
    else:
        fill = FILL_TTC
    for c in range(1, max_col + 1):
        ws.cell(row=excel_row, column=c).fill = fill


# ---------------------------------------------------------------------------
# Générateur principal
# ---------------------------------------------------------------------------

def generate_template(
    dpgf_path: str,
    output_path: str | None = None,
    lot: str = "",
    lot_name: str = "",
    keep_prices: bool = False,
) -> str:
    """
    Génère un fichier TCO modèle (.xlsx) depuis un DPGF entreprise.

    Returns:
        Chemin du fichier généré.
    """
    if not os.path.exists(dpgf_path):
        raise FileNotFoundError(f"DPGF introuvable : {dpgf_path}")

    log.info("Lecture DPGF source : %s", dpgf_path)
    df, alerts = parse_dpgf(dpgf_path)

    if df.empty:
        raise ValueError(f"DPGF vide ou non parsable. Alertes : {alerts}")

    if not lot:
        lot = _auto_detect_lot(df)
    lot_clean = lot.replace(" ", "_")

    if not lot_name:
        lot_name = os.path.splitext(os.path.basename(dpgf_path))[0].upper()

    if not output_path:
        out_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "Template_DPGF",
            "TCO_MODELE",
        )
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(
            out_dir, f"TCO {lot_clean} - {lot_name}.xlsx"
        )

    n_art = len(df[df["row_type"] == "article"])
    n_sec = len(df[df["row_type"] == "section_header"])
    n_sub = len(df[df["row_type"] == "sub_section"])
    print(f"\nDPGF source : {os.path.basename(dpgf_path)}")
    print(f"  Lot         : {lot}")
    print(f"  Structure   : {n_sec} sections | {n_sub} sous-sect. | {n_art} articles")
    if alerts:
        print(f"  Alertes DPGF ({len(alerts)}) :")
        for a in alerts[:5]:
            print(f"    [{a['type'].upper()}] {a.get('code', '')} — {a['message']}")
        if len(alerts) > 5:
            print(f"    ... et {len(alerts) - 5} autres")

    final_rows = _build_final_rows(df, lot_clean)
    log.info("Lignes finales : %d", len(final_rows))

    # --- Génération Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TCO {lot_clean}"

    # Ligne 1 : métadonnées
    ws.cell(row=1, column=1, value="LOT :").font = Font(
        name="Tahoma", bold=True, size=10
    )
    ws.cell(row=1, column=2, value=f"{lot} — {lot_name}").font = Font(
        name="Tahoma", size=10
    )

    # Ligne 2 : en-têtes (correspondance exacte avec parse_tco)
    col_defs = [
        ("A", "Code", 9.5),
        ("B", "Désignation", 56.75),
        ("C", "Qu.", 9.5),
        ("D", "U", 7.125),
        ("E", "Px U HT", 14.125),
        ("F", "Px Tot HT", 16.5),
    ]
    for col_letter, label, width in col_defs:
        col_idx = ord(col_letter) - 64
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width

    ws.cell(row=2, column=ENTETE_EXCEL_COL, value="Entete").font = FONT_HEADER
    ws.column_dimensions["M"].width = 20

    ws.row_dimensions[1].height = 14.25
    ws.row_dimensions[2].height = 14.25

    # --- Données ---
    excel_row = 3

    for row_dict in final_rows:
        rt = row_dict.get("row_type", "other")
        code = str(row_dict.get("Code", "") or "").strip()
        desig = str(row_dict.get("Désignation", "") or "").strip()

        if rt == "empty" or (not code and not desig):
            continue

        entete = row_dict.get("Entete", "") or _row_type_to_entete(rt, lot_clean)

        qu = row_dict.get("Qu.", None)
        u = str(row_dict.get("U", "") or "")
        px_u = row_dict.get("Px_U_HT", None)
        px_tot = row_dict.get("Px_Tot_HT", None)

        if not keep_prices:
            px_u = px_tot = None

        ws.cell(row=excel_row, column=1, value=code)
        ws.cell(row=excel_row, column=2, value=desig)

        if qu is not None:
            try:
                ws.cell(row=excel_row, column=3, value=float(qu))
            except (ValueError, TypeError):
                pass
        ws.cell(row=excel_row, column=4, value=u)

        if px_u is not None:
            try:
                ws.cell(row=excel_row, column=5, value=float(px_u))
            except (ValueError, TypeError):
                pass
        if px_tot is not None:
            try:
                ws.cell(row=excel_row, column=6, value=float(px_tot))
            except (ValueError, TypeError):
                pass

        ws.cell(row=excel_row, column=ENTETE_EXCEL_COL, value=entete)

        ws.cell(row=excel_row, column=3).number_format = QTY_FORMAT
        ws.cell(row=excel_row, column=5).number_format = MONEY_FORMAT
        ws.cell(row=excel_row, column=6).number_format = MONEY_FORMAT

        _style_row(ws, excel_row, rt)
        if rt == "total_line":
            _apply_total_fill(ws, excel_row, desig)
            for c in range(1, 7):
                ws.cell(row=excel_row, column=c).font = FONT_TOTAL

        indent = 0
        if rt == "article":
            indent = 2
        elif rt == "sub_section":
            indent = 1
        ws.cell(row=excel_row, column=2).alignment = Alignment(
            horizontal="left",
            indent=indent,
            wrap_text=True,
            vertical="top",
        )

        ws.row_dimensions[excel_row].height = 28.5
        excel_row += 1

    ws.freeze_panes = "A3"
    wb.save(output_path)

    n_written = excel_row - 3
    print(f"\nTemplate genere avec succes :")
    print(f"  Fichier  : {output_path}")
    print(f"  Lignes   : {n_written}")
    print("  Usage    : charger ce fichier dans l'app TCO comme 'TCO Modele'")
    return output_path


# ---------------------------------------------------------------------------
# Validation : relire via parse_tco pour vérifier la compatibilité
# ---------------------------------------------------------------------------

def validate_template(template_path: str) -> None:
    """Valide le template généré en le relisant via parse_tco()."""
    from core.parser_tco import parse_tco  # noqa: PLC0415

    print(f"\nValidation : {os.path.basename(template_path)}")
    df, _ = parse_tco(template_path)

    if df.empty:
        print("  ERREUR : parse_tco retourne un DataFrame vide !")
        return

    n_art = len(df[df["row_type"] == "article"])
    n_sec = len(df[df["row_type"] == "section_header"])
    n_sub = len(df[df["row_type"] == "sub_section"])
    n_rec = len(df[df["row_type"] == "recap"])
    n_tot = len(df[df["row_type"] == "total_line"])
    n_oth = len(df[df["row_type"] == "other"])

    print(f"  Total lignes   : {len(df)}")
    print(f"  section_header : {n_sec}")
    print(f"  sub_section    : {n_sub}")
    print(f"  article        : {n_art}")
    print(f"  recap          : {n_rec}")
    print(f"  total_line     : {n_tot}")
    if n_oth > 0:
        print(f"  [WARN] 'other' : {n_oth} lignes non reconnues")

    sections_all = set(df[df["row_type"] == "section_header"]["Code"])
    sections_with_recap = set(
        df[df["row_type"] == "recap"]["parent_code"].dropna()
    )
    missing_recap = sections_all - sections_with_recap
    if missing_recap:
        print(f"  [WARN] Sections sans recap : {sorted(missing_recap)}")
    else:
        print("  Toutes les sections ont un recap : OK")

    ok_ht = any("montant ht" in str(r).lower() for r in df["Désignation"])
    ok_tva = any("tva" in str(r).lower() for r in df["Désignation"])
    ok_ttc = any("ttc" in str(r).lower() for r in df["Désignation"])
    ht_s = "OK" if ok_ht else "MANQUANT"
    tva_s = "OK" if ok_tva else "MANQUANT"
    ttc_s = "OK" if ok_ttc else "MANQUANT"
    print(f"  Totaux HT/TVA/TTC : {ht_s} / {tva_s} / {ttc_s}")

    if n_art > 0 and n_sec > 0 and n_rec > 0:
        print("  => TEMPLATE VALIDE — pret pour l'application TCO")
    else:
        print("  => VERIFICATION MANUELLE RECOMMANDEE")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Génère un template TCO modèle depuis un DPGF entreprise",
    )
    parser.add_argument("dpgf_path", help="Chemin vers le DPGF source (.xlsx)")
    parser.add_argument("--output", default=None, help="Chemin de sortie (.xlsx)")
    parser.add_argument(
        "--lot", default="", help="Identifiant du lot (ex: Lot12)"
    )
    parser.add_argument(
        "--name", default="", help="Nom descriptif (ex: EQUIPEMENTS CUISINE)"
    )
    parser.add_argument(
        "--keep-prices",
        action="store_true",
        help="Conserver les prix du DPGF",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        default=True,
        help="Valider le template généré (défaut: True)",
    )
    parser.add_argument(
        "--no-validate",
        action="store_false",
        dest="validate",
    )

    args = parser.parse_args()

    try:
        out = generate_template(
            dpgf_path=args.dpgf_path,
            output_path=args.output,
            lot=args.lot,
            lot_name=args.name,
            keep_prices=args.keep_prices,
        )
        if args.validate:
            validate_template(out)
    except (FileNotFoundError, ValueError) as e:
        print(f"ERREUR : {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
