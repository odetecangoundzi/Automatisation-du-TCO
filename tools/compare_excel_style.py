"""
compare_excel_style.py — Comparaison automatique du style entre le fichier
de référence et un export produit par l'application.

Usage :
    python tools/compare_excel_style.py [EXPORT.xlsx]

Si EXPORT.xlsx n'est pas fourni, le script cherche le dernier fichier dans
outputs/ dont le nom commence par TCO_FINAL.
"""

from __future__ import annotations

import io
import os
import sys

# Force UTF-8 sur la console Windows (cp1252 sinon)
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import glob

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
REF_PATH = os.path.join(
    os.path.dirname(__file__), "..",
    "TCO 01 - DESAMIANTAGE - CURAGE - GO.xlsx"
)

# Largeurs de colonnes attendues (référence)
EXPECTED_WIDTHS: dict[str, float] = {
    "A": 9.5, "B": 56.75, "C": 9.5, "D": 7.125,
    "E": 14.125, "F": 16.5,
}

# Polices et tailles attendues par type de ligne (règles de la référence)
EXPECTED_FONTS = {
    "section_header": {"name": "Tahoma", "bold": True, "size": 11, "color": "AC2C18"},
    "sub_section":    {"name": "Tahoma", "bold": True, "size": 9,  "color": "314E85"},
    "article":        {"name": "Tahoma", "bold": False, "size": 9, "color": "000000"},
    "recap":          {"name": "Tahoma", "bold": True, "size": 11, "color": "000000"},
    "total_line":     {"name": "Tahoma", "bold": True, "size": 11, "color": "000000"},
}

EXPECTED_MONEY_FORMAT = r'###,###,###,##0.00\ \€;\-###,###,###,##0.00\ \€;'
EXPECTED_QTY_FORMAT   = r'###,###,###,##0.00;\-###,###,###,##0.00;'

DATA_ROW_HEIGHT = 28.5
HEADER_ROW_HEIGHT = 14.25

RESULTS: list[dict] = []


def _log(severity: str, sheet: str, location: str, prop: str, ref: str, actual: str):
    RESULTS.append({
        "severity": severity,
        "sheet": sheet,
        "location": location,
        "property": prop,
        "reference": ref,
        "actual": actual,
    })


def check_column_widths(ws_export, sheet_name: str) -> None:
    """Vérifie les largeurs des colonnes A–F."""
    for col, expected in EXPECTED_WIDTHS.items():
        actual = ws_export.column_dimensions.get(col, None)
        actual_w = actual.width if actual else None
        if actual_w is None or abs(actual_w - expected) > 0.5:
            _log(
                "Critique" if col in ("B",) else "Important",
                sheet_name, f"Col {col}", "width",
                str(expected), str(actual_w)
            )


def check_row_heights(ws_export, sheet_name: str, sample_data_rows=range(3, 20)) -> None:
    """Vérifie les hauteurs des lignes d'en-tête et de données."""
    for r in (1, 2):
        h = ws_export.row_dimensions[r].height if r in ws_export.row_dimensions else None
        if h is None or abs(h - HEADER_ROW_HEIGHT) > 1.0:
            _log("Important", sheet_name, f"Row {r}", "height",
                 str(HEADER_ROW_HEIGHT), str(h))

    for r in sample_data_rows:
        h = ws_export.row_dimensions.get(r)
        if h is None:
            _log("Important", sheet_name, f"Row {r}", "height (not set)",
                 str(DATA_ROW_HEIGHT), "default (not set)")
            break
        if h and abs(h.height - DATA_ROW_HEIGHT) > 1.0:
            _log("Important", sheet_name, f"Row {r}", "height",
                 str(DATA_ROW_HEIGHT), str(h.height))
            break


def check_freeze_panes(ws_export, sheet_name: str) -> None:
    """Vérifie le freeze panes."""
    fp = str(ws_export.freeze_panes)
    if fp != "C3":
        _log("Important", sheet_name, "freeze_panes", "freeze_panes",
             "C3", fp)


def check_number_formats(ws_export, sheet_name: str) -> None:
    """
    Vérifie que les colonnes E, F (et colonnes entreprise) utilisent
    le format monétaire de la référence.
    """
    found_money = set()
    found_qty   = set()
    for row in ws_export.iter_rows(min_row=3, max_row=min(30, ws_export.max_row)):
        for cell in row:
            fmt = cell.number_format or ""
            if fmt == EXPECTED_MONEY_FORMAT:
                found_money.add(cell.column)
            elif fmt == EXPECTED_QTY_FORMAT:
                found_qty.add(cell.column)
            elif fmt and fmt != "General" and cell.column in (3, 5, 6):
                # Colonne de prix/qté avec format différent
                col_name = get_column_letter(cell.column)
                _log(
                    "Important",
                    sheet_name,
                    f"{col_name}{cell.row}",
                    "number_format",
                    EXPECTED_MONEY_FORMAT if cell.column in (5, 6) else EXPECTED_QTY_FORMAT,
                    fmt
                )
    if not found_money:
        _log("Critique", sheet_name, "Cols E/F+", "number_format",
             EXPECTED_MONEY_FORMAT, "Aucune cellule avec ce format")
    if not found_qty:
        _log("Important", sheet_name, "Col C+", "number_format",
             EXPECTED_QTY_FORMAT, "Aucune cellule avec ce format")


def check_fonts(ws_export, sheet_name: str, sample_rows: int = 30) -> None:
    """
    Vérifie la police (Tahoma) sur les premières lignes de données.
    """
    non_tahoma = 0
    checked = 0
    for row in ws_export.iter_rows(min_row=3,
                                   max_row=min(sample_rows + 2, ws_export.max_row),
                                   min_col=1, max_col=6):
        for cell in row:
            if cell.value is None:
                continue
            checked += 1
            font_name = cell.font.name if cell.font else None
            if font_name and font_name.lower() != "tahoma":
                non_tahoma += 1
    if non_tahoma > 0:
        _log("Critique", sheet_name, f"Lignes 3–{sample_rows + 2}",
             "font.name",
             "Tahoma",
             f"{non_tahoma}/{checked} cellules ≠ Tahoma")


def check_fills(ws_export, sheet_name: str, sample_rows: int = 30) -> None:
    """
    Vérifie que les lignes de données n'ont pas de fond coloré
    (uniquement blanc FFFFFF ou transparent).
    """
    colored = 0
    checked = 0
    for row in ws_export.iter_rows(min_row=3,
                                   max_row=min(sample_rows + 2, ws_export.max_row),
                                   min_col=1, max_col=6):
        for cell in row:
            if cell.value is None:
                continue
            checked += 1
            try:
                fg = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else None
                pt = cell.fill.patternType
                if pt == "solid" and fg not in (None, "00000000", "FFFFFFFF"):
                    colored += 1
            except Exception:
                pass
    if colored > 0:
        _log("Important", sheet_name, f"Lignes 3–{sample_rows + 2}",
             "fill.fgColor",
             "FFFFFF (blanc) pour toutes les données",
             f"{colored}/{checked} cellules avec fond coloré ≠ blanc")


def check_borders(ws_export, sheet_name: str, sample_rows: int = 20) -> None:
    """Vérifie que seuls des bordures 'thin' sont utilisées (pas de medium/thick)."""
    bad_border = 0
    checked = 0
    for row in ws_export.iter_rows(min_row=3,
                                   max_row=min(sample_rows + 2, ws_export.max_row),
                                   min_col=1, max_col=6):
        for cell in row:
            b = cell.border
            for side in (b.top, b.bottom, b.left, b.right):
                if side and side.border_style and side.border_style not in ("thin", None):
                    bad_border += 1
                checked += 1
    if bad_border > 0:
        _log("Important", sheet_name, f"Lignes 3–{sample_rows + 2}",
             "border_style",
             "thin (uniquement)",
             f"{bad_border}/{checked} côtés ≠ thin")


def check_merges(ws_export, sheet_name: str) -> None:
    """Vérifie la présence d'au moins une fusion A:B sur les lignes de récap."""
    merges = [str(m) for m in ws_export.merged_cells.ranges]
    ab_merges = [m for m in merges if m.startswith("A") and ":B" in m]
    row1_merges = [m for m in merges if m.endswith("1")]
    if not ab_merges:
        _log("Mineur", sheet_name, "Merged cells A:B",
             "merged_cells", ">=1 fusion A:B (lignes recap)", "Aucune")
    if not row1_merges:
        _log("Mineur", sheet_name, "Merged cells row 1",
             "merged_cells", "Fusions groupes en-tête (ex C1:F1)", "Aucune")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_checks(export_path: str) -> None:
    print(f"\n{'='*70}")
    print("COMPARAISON STYLE EXCEL")
    print(f"Référence : {os.path.basename(REF_PATH)}")
    print(f"Export    : {os.path.basename(export_path)}")
    print(f"{'='*70}\n")

    if not os.path.exists(export_path):
        print(f"ERREUR : fichier export introuvable : {export_path}")
        sys.exit(1)

    wb_export = openpyxl.load_workbook(export_path)
    ws_export = wb_export.active
    sheet_name = ws_export.title

    # Run all checks
    check_column_widths(ws_export, sheet_name)
    check_row_heights(ws_export, sheet_name)
    check_freeze_panes(ws_export, sheet_name)
    check_number_formats(ws_export, sheet_name)
    check_fonts(ws_export, sheet_name)
    check_fills(ws_export, sheet_name)
    check_borders(ws_export, sheet_name)
    check_merges(ws_export, sheet_name)

    # Display results
    if not RESULTS:
        print("OK - Toutes les verifications de style sont conformes !\n")
        return

    severities = {"Critique": [], "Important": [], "Mineur": []}
    for r in RESULTS:
        severities.get(r["severity"], severities["Mineur"]).append(r)

    for sev, items in severities.items():
        if not items:
            continue
        emoji = {"Critique": "[CRIT]", "Important": "[WARN]", "Mineur": "[INFO]"}[sev]
        print(f"\n{emoji} {sev.upper()} ({len(items)} écart(s))")
        print("-" * 60)
        for item in items:
            print(f"  [{item['location']}] {item['property']}")
            print(f"    Référence : {item['reference']}")
            print(f"    Export    : {item['actual']}")

    total = len(RESULTS)
    crit  = len(severities["Critique"])
    impo  = len(severities["Important"])
    minor = len(severities["Mineur"])
    print(f"\n{'='*70}")
    print(f"RESUME : {total} ecart(s) -- "
          f"[CRIT] {crit}, [WARN] {impo}, [INFO] {minor}")
    if crit == 0:
        print("=> Rendu visuel correct (aucun ecart critique).")
    else:
        print("=> Verifier les ecarts critiques avant livraison.")
    print()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        export_file = sys.argv[1]
    else:
        # Chercher le dernier TCO_FINAL dans outputs/
        pattern = os.path.join(
            os.path.dirname(__file__), "..", "outputs", "TCO_FINAL*.xlsx"
        )
        files = sorted(glob.glob(pattern))
        if not files:
            # Fallback : TCO_FINAL.xlsx à la racine
            fallback = os.path.join(os.path.dirname(__file__), "..", "TCO_FINAL.xlsx")
            if os.path.exists(fallback):
                export_file = fallback
            else:
                print("Aucun fichier export trouvé. "
                      "Usage : python tools/compare_excel_style.py EXPORT.xlsx")
                sys.exit(1)
        else:
            export_file = files[-1]

    run_checks(os.path.abspath(export_file))
