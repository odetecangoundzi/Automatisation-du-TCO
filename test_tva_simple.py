import os
import openpyxl
from core.parser_tco import parse_tco
from core.parser_dpgf import parse_dpgf
from core.merger import merge_company_into_tco
from core.exporter import export_tco

def test_tva(rate):
    tco_df, meta = parse_tco(r"d:\CTO\TCO_FINAL.xlsx")
    dpgf_df, _ = parse_dpgf(r"d:\CTO\DPGF LOT 01 - DESAMIANTAGE - CURAGE - GROS OEUVRE.xlsx")
    merged, _ = merge_company_into_tco(tco_df, dpgf_df, "TEST_COMP", tva_rate=rate)
    buffer = export_tco(merged, meta, output_path=None, alerts=[], tva_rate=rate)
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    
    found = False
    for r in range(1, ws.max_row + 1):
        desig = str(ws.cell(r, 2).value).lower()
        if "tva" in desig and "ht" not in desig:
            f = str(ws.cell(r, 6).value)
            i = str(ws.cell(r, 9).value)
            if str(rate) in f and str(rate) in i:
                found = True
                print(f"RATE_MATCH_{rate}_SUCCESS")
                break
    if not found:
        print(f"RATE_MATCH_{rate}_FAILED")

if __name__ == "__main__":
    test_tva(0.055)
    test_tva(0.10)
    test_tva(0.20)
