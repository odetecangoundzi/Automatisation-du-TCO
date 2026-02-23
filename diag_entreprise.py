"""
Script de diagnostic complet des DPGF entreprise.
Sort le résultat dans un fichier UTF-8.
"""
import os
import traceback
import openpyxl
from core.parser_dpgf import parse_dpgf
from core.utils import find_header_row

ENTREPRISE_DIR = "entreprise"
OUTPUT = "diag_result.txt"

files = sorted(os.listdir(ENTREPRISE_DIR))

with open(OUTPUT, "w", encoding="utf-8") as out:
    out.write(f"=== {len(files)} fichiers trouvés dans {ENTREPRISE_DIR}/ ===\n\n")

    for fname in files:
        if not fname.endswith('.xlsx'):
            continue
        fpath = os.path.join(ENTREPRISE_DIR, fname)
        out.write(f"{'='*80}\n")
        out.write(f"FICHIER : {fname}\n")
        out.write(f"{'='*80}\n")

        # 1. Tester l'ouverture
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            ws = wb.active
            out.write(f"  Feuille active : '{ws.title}'\n")
            out.write(f"  Dimensions : {ws.max_row} lignes x {ws.max_column} colonnes\n")
        except Exception as e:
            out.write(f"  ERREUR OUVERTURE : {e}\n\n")
            continue

        # 2. Tester l'en-tête
        try:
            header_row = find_header_row(ws)
            header_cells = []
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                val = ws.cell(row=header_row, column=col_idx).value
                header_cells.append(str(val) if val else "")
            out.write(f"  En-tête trouvée ligne {header_row} : {header_cells[:7]}\n")

            col_m_val = ws.cell(row=header_row, column=13).value
            out.write(f"  Colonne M (Entete) header : '{col_m_val}'\n")

            m_count = 0
            for r in range(header_row + 1, min(header_row + 50, ws.max_row + 1)):
                if ws.cell(row=r, column=13).value:
                    m_count += 1
            out.write(f"  Colonne M remplie (50 premières lignes) : {m_count}/50\n")

        except ValueError as e:
            out.write(f"  EN-TÊTE NON TROUVÉE : {e}\n")
            out.write("  Premières lignes du fichier :\n")
            for r_idx in range(1, min(21, ws.max_row + 1)):
                a = ws.cell(row=r_idx, column=1).value
                b = ws.cell(row=r_idx, column=2).value
                c = ws.cell(row=r_idx, column=3).value
                d = ws.cell(row=r_idx, column=4).value
                e2 = ws.cell(row=r_idx, column=5).value
                f2 = ws.cell(row=r_idx, column=6).value
                out.write(f"    Ligne {r_idx}: A='{a}' | B='{b}' | C='{c}' | D='{d}' | E='{e2}' | F='{f2}'\n")
            wb.close()
            out.write("\n")
            continue

        wb.close()

        # 3. Tester le parsing complet
        try:
            dpgf_df, alerts = parse_dpgf(fpath)

            row_types = dpgf_df['row_type'].value_counts().to_dict()
            n_articles = row_types.get('article', 0)
            n_other = row_types.get('other', 0)
            n_total = len(dpgf_df)

            out.write(f"  Total lignes parsées : {n_total}\n")
            out.write(f"  Types de lignes : {row_types}\n")
            out.write(f"  Articles (fusionnables) : {n_articles}\n")
            out.write(f"  Alertes : {len(alerts)}\n")

            articles = dpgf_df[dpgf_df['row_type'] == 'article']
            if n_articles > 0:
                codes_sample = articles['Code'].head(5).tolist()
                out.write(f"  Exemples de codes article : {codes_sample}\n")
                with_prices = articles[articles['Px_Tot_HT'] > 0]
                out.write(f"  Articles avec prix > 0 : {len(with_prices)}/{n_articles}\n")
            else:
                out.write(f"  *** AUCUN ARTICLE DÉTECTÉ ***\n")
                others_with_price = dpgf_df[(dpgf_df['row_type'] == 'other') & (dpgf_df['Px_Tot_HT'] > 0)]
                if len(others_with_price) > 0:
                    out.write(f"  -> {len(others_with_price)} lignes 'other' ont un prix\n")
                    for _, row in others_with_price.head(3).iterrows():
                        out.write(f"    Code='{row['Code']}' Desig='{row['Désignation'][:50]}' PxTot={row['Px_Tot_HT']}\n")

            if n_articles == 0 and n_other > 0:
                out.write(f"  PROBLÈME : Aucun article mais {n_other} lignes 'other'\n")

        except Exception as e:
            out.write(f"  ERREUR PARSING : {e}\n")
            out.write(traceback.format_exc())

        out.write("\n")

    out.write("=== FIN DU DIAGNOSTIC ===\n")

print(f"Résultat écrit dans {OUTPUT}")
