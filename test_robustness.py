import os
import traceback
import pandas as pd
from core.parser_dpgf import parse_dpgf
from core.parser_tco import parse_tco
import openpyxl

print("=== DEBUT DES TESTS DE ROBUSTESSE EXCEL ===")

def run_robustness_test(filename, description, setup_fn):
    print(f"\n--- Test: {description} ---")
    try:
        setup_fn(filename)
        print("  Fichier cree.")
        df, alerts = parse_dpgf(filename)
        print(f"  [OK] Parse {len(df)} lignes, {len(alerts)} alertes.")
        # On peut ajouter des assertions si besoin
        assert isinstance(df, pd.DataFrame)
    except Exception as e:
        print(f"  [ERROR] {type(e).__name__} - {e}")
        raise
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except OSError:
                pass

# 1. Fichier complètement vide
def create_empty(fname):
    wb = openpyxl.Workbook()
    wb.save(fname)

def test_empty_file():
    run_robustness_test("test_empty.xlsx", "Fichier Excel vide", create_empty)

# 2. Fichier avec des données mais pas d'en-tête
def create_no_header(fname):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Coucou", "Ceci", "N'est", "Pas", "Un", "DPGF"])
    wb.save(fname)

def test_no_header():
    run_robustness_test("test_no_header.xlsx", "Fichier sans en-tete standard", create_no_header)

# 3. Fichier avec en-tête mais sans lignes de données
def create_only_header(fname):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Code", "Designation", "Qu.", "U", "Px U", "Px Tot"])
    wb.save(fname)

def test_only_header():
    run_robustness_test("test_only_header.xlsx", "Fichier avec entete mais sans donnees", create_only_header)

# 4. Fichier avec en-tête mais colonnes décalées
def create_shifted_header(fname):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["", "", "", "Code", "Designation", "Qu.", "U", "Px U", "Px Tot"])
    ws.append(["", "", "", "01.1", "Truc", 1, "m2", 10, 10])
    wb.save(fname)

def test_shifted_header():
    run_robustness_test("test_shifted.xlsx", "Colonnes decalees horizontalement", create_shifted_header)

# 5. Fichier avec lignes de données tronquées
def create_truncated(fname):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Code", "Designation", "Qu.", "U", "Px U", "Px Tot"]) 
    ws.append(["01.1", "Article sans prix"]) 
    wb.save(fname)

def test_truncated_file():
    run_robustness_test("test_truncated.xlsx", "Fichier sans les colonnes de prix/quantite", create_truncated)

if __name__ == "__main__":
    # Permet de toujours lancer le script à la main
    import pytest
    pytest.main([__file__])
