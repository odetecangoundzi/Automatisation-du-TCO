"""
parser_dpgf.py — Normalisation et extraction des annotations des DPGF entreprise.

Gère les textes dans les cellules numériques (SANS OBJET, COMPRIS, nc, P-M),
extrait les annotations en colonne Commentaire, et détecte les erreurs.
Utilise la colonne Entete (col M) pour classifier chaque ligne.
"""

import pandas as pd
import openpyxl
import re


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

KEYWORDS = {
    "sans objet": "so",
    "compris": "compris",
    "nc": "nc",
    "p-m": "pm",
    "inclus": "inclus",
    "néant": "néant",
    "so": "so",
    "pm": "pm",
}

TOTAL_TOLERANCE = 0.02


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_header_row(ws):
    """Trouve la ligne d'en-tête dans un DPGF entreprise."""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        a_val = ws.cell(row=row_idx, column=1).value
        b_val = ws.cell(row=row_idx, column=2).value
        if (
            a_val and str(a_val).strip().lower() == "code"
            and b_val and "signation" in str(b_val).strip().lower()
        ):
            return row_idx
    raise ValueError("Impossible de trouver la ligne d'en-tête du DPGF.")


def _classify_row(code_str, desig_str, entete_str):
    """
    Classifie une ligne du DPGF selon la colonne Entete.
    Même logique que parser_tco._classify_row.
    """
    ent = entete_str
    if "RecapBord" in ent:
        return "recap_summary"
    if "LignesTot" in ent:
        return "total_line"
    if "Bord" in ent and "Recap" in ent:
        return "recap"
    if ent.startswith("Bd_") and "Bord" in ent:
        return "section_header"
    if "_Niv1" in ent or "_Niv2" in ent:
        return "sub_section"
    if "_Art" in ent:
        return "article"
    if code_str.lower().startswith("total"):
        return "total_text"
    if not code_str and not desig_str:
        return "empty"
    return "other"


def _clean_numeric(value):
    """
    Nettoie une valeur potentiellement numérique.
    Retourne (nombre_float, texte_annotation).
    """
    if value is None:
        return 0.0, ""

    if isinstance(value, (int, float)):
        return float(value), ""

    text = str(value).strip()
    if not text:
        return 0.0, ""

    # Mot-clé connu
    text_lower = text.lower().strip()
    for keyword, abbrev in KEYWORDS.items():
        if keyword in text_lower:
            return 0.0, abbrev

    # Format français → nombre
    cleaned = text.replace(" ", "").replace("\u00a0", "").replace(",", ".")

    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if match:
        number = float(match.group())
        remaining = re.sub(r"-?\d+(?:\.\d+)?", "", cleaned).strip()
        remaining = remaining.strip("()[]{}/ ")
        return number, remaining

    return 0.0, text


def _check_total_coherence(qu_val, pu_val, total_val, row_idx, code):
    """Vérifie que Qu × PU ≈ Total."""
    if qu_val and pu_val and total_val:
        try:
            expected = float(qu_val) * float(pu_val)
            actual = float(total_val)
            if actual != 0 and abs(expected - actual) > TOTAL_TOLERANCE:
                return {
                    "type": "error",
                    "color": "red",
                    "row": row_idx,
                    "code": code,
                    "message": (
                        f"Total incohérent : {qu_val} × {pu_val} = "
                        f"{expected:.2f} ≠ {actual:.2f}"
                    ),
                }
        except (ValueError, TypeError):
            pass
    return None


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_dpgf(filepath):
    """
    Lit et normalise un fichier DPGF entreprise (.xlsx).

    Retourne :
        dpgf_df : DataFrame normalisé
        alerts  : liste d'alertes
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    alerts = []

    rows = []
    current_section_code = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        code_raw = ws.cell(row=row_idx, column=1).value
        desig_raw = ws.cell(row=row_idx, column=2).value
        cc_raw = ws.cell(row=row_idx, column=3).value
        u = ws.cell(row=row_idx, column=4).value
        px_u_raw = ws.cell(row=row_idx, column=5).value
        px_tot_raw = ws.cell(row=row_idx, column=6).value
        entete = ws.cell(row=row_idx, column=13).value

        code_str = str(code_raw).strip() if code_raw else ""
        desig_str = str(desig_raw).strip() if desig_raw else ""
        ent_str = str(entete).strip() if entete else ""

        row_type = _classify_row(code_str, desig_str, ent_str)

        # Suivre la section parente
        if row_type == "section_header":
            current_section_code = code_str

        parent_code = ""
        if row_type == "recap":
            parent_code = current_section_code

        # Normaliser les valeurs numériques uniquement pour articles
        if row_type == "article":
            qu_val, qu_comment = _clean_numeric(cc_raw)
            pu_val, pu_comment = _clean_numeric(px_u_raw)
            tot_val, tot_comment = _clean_numeric(px_tot_raw)
        elif row_type == "sub_section" and cc_raw is not None:
            # Sous-sections avec quantités renseignées
            qu_val, qu_comment = _clean_numeric(cc_raw)
            pu_val, pu_comment = _clean_numeric(px_u_raw)
            tot_val, tot_comment = _clean_numeric(px_tot_raw)
        else:
            qu_val = cc_raw if isinstance(cc_raw, (int, float)) else 0.0
            pu_val = px_u_raw if isinstance(px_u_raw, (int, float)) else 0.0
            tot_val = px_tot_raw if isinstance(px_tot_raw, (int, float)) else 0.0
            qu_comment = pu_comment = tot_comment = ""

        # Construire le commentaire consolidé
        comments = [c for c in [qu_comment, pu_comment, tot_comment] if c]
        commentaire = "; ".join(comments) if comments else ""

        # Alertes pour articles
        if row_type == "article" and code_str:
            if qu_comment or pu_comment or tot_comment:
                kw_found = any(
                    c.lower() in KEYWORDS or c.lower() in KEYWORDS.values()
                    for c in [qu_comment, pu_comment, tot_comment] if c
                )
                if kw_found:
                    alerts.append({
                        "type": "info", "color": "blue",
                        "row": row_idx, "code": code_str,
                        "message": f"Mot-clé détecté : {commentaire}",
                    })
                else:
                    alerts.append({
                        "type": "warning", "color": "yellow",
                        "row": row_idx, "code": code_str,
                        "message": f"Texte dans champ numérique : {commentaire}",
                    })

            alert = _check_total_coherence(
                qu_val, pu_val, tot_val, row_idx, code_str
            )
            if alert:
                alerts.append(alert)

        rows.append({
            "Code": code_str,
            "Désignation": desig_str,
            "Qu.": qu_val,
            "U": str(u).strip() if u else "",
            "Px_U_HT": pu_val,
            "Px_Tot_HT": tot_val,
            "Commentaire": commentaire,
            "Entete": ent_str,
            "row_type": row_type,
            "original_row": row_idx,
            "parent_code": parent_code,
        })

    dpgf_df = pd.DataFrame(rows)
    wb.close()
    return dpgf_df, alerts
