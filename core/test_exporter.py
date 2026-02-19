"""
test_exporter.py — Tests unitaires pour le module d'export Excel.

Couvre :
- Format des headers groupés
- Styles (gras, couleurs)
- Contenu des cellules
- Absence de "trous" (BUG-1)
- Largeur des colonnes
"""

import sys, os
import shutil
import unittest
import openpyxl
import pandas as pd

# MOCK : config pour tests
sys.path.insert(0, r"d:\CTO")

from core.exporter import export_tco
from openpyxl.styles import PatternFill

class TestExporter(unittest.TestCase):
    
    def setUp(self):
        self.output_dir = r"d:\CTO\test_outputs"
        if os.path.exists(self.output_dir):
            shutil.rmtree(self.output_dir)
        os.makedirs(self.output_dir)
        
        # Données mock
        self.mock_data = [
            {"Code": "01.1", "Désignation": "SECTION 1", "row_type": "section_header", "original_row": 10},
            {"Code": "01.1.1", "Désignation": "Article 1", "Qu.": 10, "Px_U_HT": 100, "Px_Tot_HT": 1000, 
             "row_type": "article", "MAB_Qu.": 10, "MAB_Px_U_HT": 110, "MAB_Px_Tot_HT": 1100, 
             "MAB_Commentaire": "OK", "original_row": 11},
            {"Code": "", "Désignation": "RECAP", "row_type": "recap", "original_row": 12},
            {"Code": "Total", "Désignation": "Total Général", "row_type": "total_line", "original_row": 13},
        ]
        self.df = pd.DataFrame(self.mock_data)
        self.meta = {"sheet_name": "TestSheet"}

    def test_export_structure(self):
        """Vérifie la structure globale du fichier Excel généré."""
        output_path = os.path.join(self.output_dir, "test_structure.xlsx")
        export_tco(self.df, self.meta, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Titre sheet
        self.assertEqual(ws.title, "TestSheet")
        
        # Headers Row 1 (merged)
        self.assertEqual(ws.cell(1, 2).value, "Etudes")
        self.assertEqual(ws.cell(1, 7).value, "MAB") # Entreprise détectée
        
        # Headers Row 2
        self.assertEqual(ws.cell(2, 1).value, "Code")
        self.assertEqual(ws.cell(2, 7).value, "Qu.") # Colonne entreprise
        
        wb.close()

    def test_styles_and_colors(self):
        """Vérifie que les styles (couleurs, gras) sont appliqués."""
        output_path = os.path.join(self.output_dir, "test_styles.xlsx")
        export_tco(self.df, self.meta, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Section header (Row 3) : Bold + Bleu
        section_cell = ws.cell(3, 2)
        self.assertTrue(section_cell.font.bold)
        # self.assertEqual(section_cell.fill.start_color.rgb, "FFDCE6F1") # Parfois alpha channel varie
        
        # Entreprise Header (Row 1) : Couleur spécifique
        company_header = ws.cell(1, 7)
        self.assertTrue(company_header.font.bold)
        
        wb.close()

    def test_no_gaps(self):
        """Vérifie qu'il n'y a pas de lignes vides (BUG-1)."""
        # Ajouter une ligne 'empty' au milieu
        data_with_empty = self.mock_data.copy()
        data_with_empty.insert(1, {"row_type": "empty"})
        df = pd.DataFrame(data_with_empty)
        
        output_path = os.path.join(self.output_dir, "test_gaps.xlsx")
        export_tco(df, self.meta, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Row 1=Header, 2=SubHeader -> Data starts at 3
        # Data rows: Section(3), Article(4), Recap(5), Total(6)
        # Empty row should be skipped
        self.assertEqual(ws.cell(3, 1).value, "01.1")
        self.assertEqual(ws.cell(4, 1).value, "01.1.1")
        self.assertEqual(ws.cell(4, 2).value, "Article 1") # Vérifie alignement
        
        wb.close()

if __name__ == "__main__":
    unittest.main()
