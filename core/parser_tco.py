"""
parser_tco.py — Lecture et validation du TCO modèle (fichier DPGF LOT).

Détecte dynamiquement la ligne d'en-tête, extrait les colonnes
Code/Désignation/Qu./U./Px U./Px tot. et les métadonnées du projet.
Classifie chaque ligne selon son type (section, recap, article, total, etc.)
en se basant sur la colonne Entete (col M).
"""

import pandas as pd
import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_header_row(ws):
    """
    Parcourt les lignes pour trouver celle contenant 'Code' en col A
    et 'Désignation' en col B. Retourne le numéro de ligne (1-indexed).
    """
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        a_val = ws.cell(row=row_idx, column=1).value
        b_val = ws.cell(row=row_idx, column=2).value
        if (
            a_val and str(a_val).strip().lower() == "code"
            and b_val and "signation" in str(b_val).strip().lower()
        ):
            return row_idx
    raise ValueError(
        "Impossible de trouver la ligne d'en-tête (Code | Désignation)."
    )


def _extract_project_info(ws, header_row):
    """
    Extrait les informations du projet situées au-dessus de la ligne d'en-tête.
    """
    info = {}
    labels = ["region", "projet", "adresse", "phase", "lot"]
    idx = 0
    for row_idx in range(1, header_row):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip():
            if idx < len(labels):
                info[labels[idx]] = str(val).strip()
                idx += 1
    return info


def _classify_row(code_str, desig_str, entete_str):
    """
    Classifie une ligne selon les métadonnées Entete :
      - section_header : Bd_xxx_Bord (section principale avec Code)
      - recap          : Bord_xxx_Recap (totalisation, Code souvent vide)
      - recap_summary  : RecapBord_xxx (table récap en fin de fichier)
      - sub_section    : Ouv_xxx_Niv1 / Ouv_xxx_Niv2
      - article        : Ouv_xxx_Art (ligne de détail avec prix)
      - total_line     : LignesTot_xxx (Montant HT, TVA, TTC)
      - total_text     : ligne dont le code commence par 'Total'
      - empty          : rien
      - other          : tout le reste
    """
    ent = entete_str

    if "RecapBord" in ent:
        return "recap_summary"
    if "LignesTot" in ent:
        return "total_line"
    if "Bord" in ent and "Recap" in ent:
        return "recap"
    if ent.startswith("Bd_") and "Bord" in ent:
        return "section_header"
    if "_Niv1" in ent or "_Niv2" in ent:
        return "sub_section"
    if "_Art" in ent:
        return "article"
    if code_str.lower().startswith("total"):
        return "total_text"
    if not code_str and not desig_str:
        return "empty"
    return "other"


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_tco(filepath):
    """
    Lit un fichier TCO modèle (DPGF LOT .xlsx).

    Retourne :
        tco_df : DataFrame avec colonnes
            [Code, Désignation, Qu., U, Px_U_HT, Px_Tot_HT,
             Entete, row_type, original_row, parent_code]
        meta   : dict contenant project_info, header_row, sheet_name
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    sheet_name = ws.title

    header_row = _find_header_row(ws)
    project_info = _extract_project_info(ws, header_row)

    rows = []
    current_section_code = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        code_raw = ws.cell(row=row_idx, column=1).value
        desig_raw = ws.cell(row=row_idx, column=2).value
        qu = ws.cell(row=row_idx, column=3).value
        u = ws.cell(row=row_idx, column=4).value
        px_u = ws.cell(row=row_idx, column=5).value
        px_tot = ws.cell(row=row_idx, column=6).value
        entete = ws.cell(row=row_idx, column=13).value  # col M

        code_str = str(code_raw).strip() if code_raw else ""
        desig_str = str(desig_raw).strip() if desig_raw else ""
        ent_str = str(entete).strip() if entete else ""

        row_type = _classify_row(code_str, desig_str, ent_str)

        # Suivre la section parente pour les recap
        if row_type == "section_header":
            current_section_code = code_str

        parent_code = ""
        if row_type == "recap":
            parent_code = current_section_code

        rows.append({
            "Code": code_str,
            "Désignation": desig_str,
            "Qu.": qu,
            "U": str(u).strip() if u else "",
            "Px_U_HT": px_u,
            "Px_Tot_HT": px_tot,
            "Entete": ent_str,
            "row_type": row_type,
            "original_row": row_idx,
            "parent_code": parent_code,
        })

    tco_df = pd.DataFrame(rows)

    meta = {
        "project_info": project_info,
        "header_row": header_row,
        "sheet_name": sheet_name,
        "filepath": filepath,
    }

    wb.close()
    return tco_df, meta
