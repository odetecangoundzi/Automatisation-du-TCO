"""
exporter.py — Export du TCO fusionné en fichier Excel formaté.

Génère un fichier .xlsx avec :
- Headers groupés par entreprise (merged cells)
- Mise en forme (gras, freeze pane, largeur auto)
- Coloration selon alertes
- Lignes section_header et recap mises en évidence
- Support export via BytesIO (pas de sauvegarde disque obligatoire)
"""

from __future__ import annotations

import io
import re
from copy import copy as _copy
from decimal import Decimal

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from logger import get_logger

log = get_logger(__name__)


# ---------------------------------------------------------------------------
# Regex pré-compilées — évite la recompilation à chaque ligne exportée
# (la boucle principale iterrows() appelle ces patterns sur chaque ligne)
# ---------------------------------------------------------------------------

_RE_MONTANT_HT = re.compile(r"montant\s+ht")
_RE_TVA_ONLY = re.compile(r"\btva\b")
_RE_HT_ONLY = re.compile(r"\bht\b")
_RE_MONTANT_TTC = re.compile(r"montant\s+ttc|(?<!\w)ttc(?!\w)")
_RE_ALT_ARTICLE = re.compile(r"\d[A-Z]+$")
_RE_LOT_NUM = re.compile(r"\b(\d{2})\b")


def _clean_val(v):
    """Convertit pour openpyxl en évitant de transformer le texte en None.
    - Transforme Decimal en float.
    - Laisse les chaines ("NC", "INCLUS") telles quelles.
    - Retourne None pour les nan/viudes.
    """
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, float) and v != v:  # NaN check
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None

    try:
        f = float(v)
        return None if f != f else f
    except (TypeError, ValueError):
        return v


# ---------------------------------------------------------------------------
# Constantes de style
# ---------------------------------------------------------------------------

FONT_HEADER = Font(name="Tahoma", bold=True, size=10, color="FFFFFF")
FONT_HEADER_COMPANY = Font(name="Tahoma", bold=True, size=10, color="FFFFFF")
FONT_SECTION = Font(name="Tahoma", bold=True, size=11, color="AC2C18")  # rouge foncé — référence
FONT_RECAP = Font(name="Tahoma", bold=True, size=11, color="000000")  # noir gras
FONT_RECAP_HEADER_LARGE = Font(
    name="Tahoma", bold=True, size=13, color="FFFFFF"
)  # blanc gras grand — bandeau récap
FONT_TOTAL = Font(name="Tahoma", bold=True, size=11, color="000000")  # noir gras
FONT_GRAND_TOTAL = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")  # blanc sur fond foncé
FONT_DATA = Font(name="Tahoma", size=9, color="000000")
FONT_SUB_SECTION = Font(name="Tahoma", bold=True, size=9, color="314E85")  # bleu foncé — référence

FILL_HEADER = PatternFill(start_color="1B2A47", end_color="1B2A47", fill_type="solid") # Bleu nuit corporate
FILL_COMPANY_COLORS = [
    PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid"),
    PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
]

# Teintes claires (≈15 % opacité) des couleurs entreprises.
# Appliquées aux colonnes de chaque entreprise dans les lignes recap_summary
# pour créer un lien visuel entre en-têtes et totaux du récapitulatif.
FILL_COMPANY_TINTS = [
    PatternFill(
        start_color="D9E8D1", end_color="D9E8D1", fill_type="solid"
    ),  # vert clair  (548235)
    PatternFill(
        start_color="FFF0CC", end_color="FFF0CC", fill_type="solid"
    ),  # or clair    (BF8F00)
    PatternFill(
        start_color="F5DDD5", end_color="F5DDD5", fill_type="solid"
    ),  # brun clair  (843C0C)
    PatternFill(
        start_color="EAD9F5", end_color="EAD9F5", fill_type="solid"
    ),  # violet clair(7030A0)
    PatternFill(
        start_color="FAD9D9", end_color="FAD9D9", fill_type="solid"
    ),  # rouge clair  (C00000)
]

# Lignes de données : fond blanc pur (conforme référence — hiérarchie via couleur police)
# Format ARGB 8 chars : "FFFFFFFF" = blanc opaque — correspond à fgColor.rgb de la référence
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_SECTION = FILL_WHITE
FILL_RECAP = PatternFill(  # bleu moyen — sous-total de section, clairement distinct des articles
    start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"
)
FILL_RECAP_SUMMARY = PatternFill(  # bleu clair acier
    start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
)
FILL_TOTAL_LINE = FILL_WHITE
FILL_ZEBRA_EVEN = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid") # Gris perle très discret pour lignes paires
FILL_RECAP_HEADER = PatternFill(  # marine foncé — bandeau titre récapitulatif
    start_color="17375E", end_color="17375E", fill_type="solid"
)
FILL_RECAP_SEPARATOR = PatternFill(  # gris anthracite — ligne de séparation avant récap
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
FILL_SUB_SECTION = FILL_WHITE

# Titres principaux (sub_section sans prix = ex : BATIMENT F)
FONT_MAIN_TITLE = Font(name="Tahoma", bold=True, size=11, color="314E85")  # bleu foncé ref
FILL_MAIN_TITLE = FILL_WHITE

# Totaux généraux — fond sombre + texte blanc (FONT_GRAND_TOTAL)
FILL_MONTANT_HT = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
)  # bleu foncé
FILL_TVA = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # bleu moyen
FILL_MONTANT_TTC = PatternFill(
    start_color="0D2137", end_color="0D2137", fill_type="solid"
)  # bleu très foncé

FILL_ERROR = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")     # Rouge pastel
FILL_WARNING = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")   # Jaune pastel
FILL_NOTE = PatternFill(start_color="FDF6E3", end_color="FDF6E3", fill_type="solid")
FILL_INFO = PatternFill(start_color="E6F2F9", end_color="E6F2F9", fill_type="solid")

FONT_RECAP_HEADER = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")

# Articles alternatifs (code suffixé "A", "AA"…) — orange, conforme modèles de référence
FONT_ALTERNATIVE = Font(name="Tahoma", bold=True, size=9, color="DC9329")

# Récap de section imbriquée (sous-total intermédiaire) — taille 8pt comme la référence
FONT_RECAP_SUB = Font(name="Tahoma", bold=True, size=8, color="000000")

# --- STYLES EXCLUSIFS POUR SECTIONS DYNAMIQUES (CLASSES) ---
FONT_OPT_CLASS = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")
FILL_OPT_CLASS = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Violet

FONT_99_CLASS = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")
FILL_99_CLASS = PatternFill(start_color="5B5B5B", end_color="5B5B5B", fill_type="solid")  # Gris foncé

FONT_SANS_CODE_CLASS = Font(
    name="Tahoma", bold=True, size=11, color="FFFFFF"
)
FILL_SANS_CODE_CLASS = PatternFill(
    start_color="833C00", end_color="833C00", fill_type="solid"
)  # Brun

FILL_ADDED = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
)  # Blanc (ligne ajoutée par l'entreprise)

# ---------------------------------------------------------------------------
# Couleurs d'onglet par numéro de lot (palette cyclique — supporte tous les lots)
# ---------------------------------------------------------------------------
_LOT_PALETTE: list[str] = [
    "548235",  # 01 vert
    "843C0C",  # 02 brun
    "1F4E79",  # 03 bleu foncé
    "7030A0",  # 04 violet
    "2E75B6",  # 05 bleu moyen
    "BF8F00",  # 06 or
    "C00000",  # 07 rouge
    "375623",  # 08 vert foncé
    "984807",  # 09 orange foncé
    "2E4057",  # 10 ardoise
    "5C4827",  # 11 brun foncé
    "1C6E52",  # 12 vert émeraude
]

# Correspondances sémantiques historiques (lots 01-06, 10-12)
_SEMANTIC_COLORS: dict[str, str] = {
    "01": "548235",
    "02": "843C0C",
    "03": "1F4E79",
    "04": "7030A0",
    "05": "2E75B6",
    "06": "BF8F00",
    "10": "C00000",
    "11": "375623",
    "12": "984807",
}


def _get_lot_tab_color(lot_num: str) -> str:
    """Retourne la couleur d'onglet pour un numéro de lot quelconque.

    Les lots 01-06 et 10-12 conservent leurs couleurs historiques.
    Les autres lots utilisent la palette cyclique (_LOT_PALETTE).
    """
    if lot_num in _SEMANTIC_COLORS:
        return _SEMANTIC_COLORS[lot_num]
    try:
        return _LOT_PALETTE[(int(lot_num) - 1) % len(_LOT_PALETTE)]
    except (ValueError, TypeError):
        return "2F5496"  # bleu par défaut


THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
THICK_RIGHT_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="medium", color="7A7A7A"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Bordure épaisse — utilisée pour encadrer les lignes du récapitulatif
MEDIUM_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
# Bordure top épaisse + rest thin — délimite la 1ère ligne recap_summary
RECAP_TOP_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="thin"),
)
# Bordure bottom épaisse — délimite la dernière ligne recap_summary
RECAP_BOTTOM_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="thin"),
    bottom=Side(style="medium"),
)

# Formats numériques — format exact de la référence
MONEY_FORMAT = r"###,###,###,##0.00\ \€;\-###,###,###,##0.00\ \€;"
QTY_FORMAT = r"###,###,###,##0.00;\-###,###,###,##0.00;"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _detect_companies(df: pd.DataFrame) -> list[str]:
    """Détecte les noms d'entreprises à partir des colonnes _Qu."""
    companies, seen = [], set()
    for col in df.columns:
        if col.endswith("_Qu."):
            name = col[:-4]
            if name not in seen:
                companies.append(name)
                seen.add(name)
    return companies


def _auto_width(ws, min_width: int = 8, max_width: int = 40) -> None:
    """Ajuste la largeur des colonnes automatiquement."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                # Truncate string representation to check length, avoid huge cells
                val_str = str(cell.value)
                line_len = (
                    max(len(line) for line in val_str.split("\n"))
                    if "\n" in val_str
                    else len(val_str)
                )
                max_len = max(max_len, min(line_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _get_alert_fill(color: str) -> PatternFill | None:
    return {"red": FILL_ERROR, "orange": FILL_WARNING, "yellow": FILL_NOTE, "blue": FILL_INFO}.get(
        color
    )


def _get_row_style(row_type: str) -> tuple[Font, PatternFill | None]:
    return {
        "section_header": (FONT_SECTION, FILL_SECTION),
        "recap": (FONT_RECAP, FILL_RECAP),
        "recap_summary": (FONT_RECAP, FILL_RECAP_SUMMARY),
        "total_line": (FONT_TOTAL, FILL_TOTAL_LINE),
        "sub_section": (FONT_SUB_SECTION, FILL_SUB_SECTION),
        "class_opt": (FONT_OPT_CLASS, FILL_OPT_CLASS),
        "class_99": (FONT_99_CLASS, FILL_99_CLASS),
        "class_sans_code": (FONT_SANS_CODE_CLASS, FILL_SANS_CODE_CLASS),
    }.get(row_type, (FONT_DATA, None))


def fix_freeze_panes(ws, header_rows: int = 2, frozen_cols: int = 0) -> None:
    """
    Garantit que le freeze panes est positionné à la cellule ancre correcte.
    header_rows=2, frozen_cols=0  →  ancre A3
    (lignes 1-2 figées + 0 colonnes figées).
    """
    anchor = f"{get_column_letter(frozen_cols + 1)}{header_rows + 1}"
    ws.freeze_panes = anchor


def fix_merged_cells_crossing_freeze(
    ws,
    header_rows: int = 2,
    frozen_cols: int = 2,
) -> None:
    """
    Supprime toute fusion qui traverse la frontière de freeze panes.
    Les fusions horizontales sont remplacées par centerContinuous pour
    conserver l'effet visuel sans provoquer de chevauchement au scroll.
    """
    to_process = []
    for mr in list(ws.merged_cells.ranges):
        crosses_col = mr.min_col <= frozen_cols < mr.max_col
        crosses_row = mr.min_row <= header_rows < mr.max_row
        if not (crosses_col or crosses_row):
            continue
        pivot = ws.cell(row=mr.min_row, column=mr.min_col)
        to_process.append(
            (
                mr.coord,
                mr.min_row,
                mr.min_col,
                mr.max_row,
                mr.max_col,
                pivot.value,
                _copy(pivot.font) if pivot.font else None,
                _copy(pivot.fill) if pivot.fill else None,
            )
        )

    for coord, min_row, min_col, max_row, max_col, val, fnt, fll in to_process:
        ws.unmerge_cells(coord)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if fnt:
                    cell.font = fnt
                if fll:
                    cell.fill = fll
                if max_row == min_row:  # fusion horizontale → center across
                    cell.alignment = Alignment(horizontal="centerContinuous")
        ws.cell(row=min_row, column=min_col).value = val
        log.debug("Fusion crossing freeze corrigée : %s", coord)


def prevent_text_overflow(
    ws,
    min_row: int = 3,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> None:
    """
    Garantit qu'aucune cellule du tableau n'est transparente (fill_type=None).
    Un fill blanc sur les cellules vides empêche le texte adjacent de déborder
    horizontalement pendant le scroll (effet "spill" Excel).
    """
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            ft = cell.fill.fill_type if cell.fill else None
            if ft is None or ft == "none":
                cell.fill = FILL_WHITE


def _rows_to_sum_formula(col: str, rows: list[int]) -> str:
    """
    Convertit une liste de numéros de lignes Excel en formule =SUM() avec plages.

    Exemple : [3,4,5,7,8] → '=SUM(F3:F5,F7:F8)'
    Jamais d'énumération de cellules individuelles : toujours des plages contiguës.
    """
    if not rows:
        return "0"
    sorted_rows = sorted(set(rows))
    parts: list[str] = []
    start = end = sorted_rows[0]
    for r in sorted_rows[1:]:
        if r == end + 1:
            end = r
        else:
            parts.append(f"{col}{start}:{col}{end}" if start != end else f"{col}{start}")
            start = end = r
    parts.append(f"{col}{start}:{col}{end}" if start != end else f"{col}{start}")
    return "=SUM(" + ",".join(parts) + ")"


# ---------------------------------------------------------------------------
# Main exporter
# ---------------------------------------------------------------------------


def export_tco(
    merged_df: pd.DataFrame,
    meta: dict,
    output_path: str | None = None,
    alerts: list[dict] | None = None,
    tva_rate: float = 0.20,
    comparatif_mode: bool = False,
) -> str | io.BytesIO:
    """
    Exporte le TCO fusionné en fichier Excel formaté.

    comparatif_mode=True : pas de DPGF estimation — les colonnes C-F (Qu./U/Px U/Px Tot)
    sont masquées dans l'export, seules les colonnes entreprises sont visibles.
    """
    if alerts is None:
        alerts = []

    log.info("Début export Excel. Lignes=%d", len(merged_df))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta.get("sheet_name", "TCO Final")

    # --- Identification du lot : numéro + couleur d'onglet ---
    _lot_raw = ((meta.get("project_info") or {}).get("lot", "") or "").strip()
    _lot_match = _RE_LOT_NUM.search(_lot_raw)
    _lot_num = _lot_match.group(1) if _lot_match else ""
    _tab_color = _get_lot_tab_color(_lot_num) if _lot_num else "2F5496"
    ws.sheet_properties.tabColor = _tab_color

    # Affichage du récapitulatif AU-DESSUS du détail (plus intuitif pour les TCO)
    ws.sheet_properties.outlinePr.summaryBelow = False
    log.debug("Lot détecté : '%s' → tab color #%s", _lot_num or "?", _tab_color)

    companies = _detect_companies(merged_df)
    log.debug("Entreprises détectées : %s", companies)

    # --- ROW 1-4 : Metadata (MOA, MOE, etc.) ---
    project_info = meta.get("project_info", {})
    # Access metadata potentially passed down from the active project via the caller
    moa = project_info.get("moa", "")
    moe = project_info.get("moe", "")
    devise = project_info.get("devise", "€")

    current_row = 1
    if moa or moe:
        ws.cell(row=current_row, column=1, value="Maître d'Ouvrage :").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=moa)

        # --- Légende Visuelle (Premium) ---
        _l_col = 8 # Colonne H
        ws.cell(row=current_row, column=_l_col, value="Légende Analyse :").font = Font(bold=True, italic=True, size=9)

        _best = ws.cell(row=current_row, column=_l_col + 1, value="● Meilleur Prix")
        _best.font = Font(color="38761D", bold=True, size=9)

        current_row += 1
        ws.cell(row=current_row, column=1, value="Maître d'Œuvre :").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=moe)

        _nc = ws.cell(row=current_row, column=_l_col + 1, value="● NC : Non Chiffré")
        _nc.font = Font(color="999999", italic=True, size=8)

        current_row += 1
        ws.cell(row=current_row, column=1, value="Devise :").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=devise)

        _hint = ws.cell(row=current_row, column=_l_col + 1, value="💡 Astuce : Utilisez les boutons [1] [2] à gauche pour plier les détails")
        _hint.font = Font(color="2F5496", size=8, italic=True)

        current_row += 2  # Add a blank line

    header_row_1 = current_row
    header_row_2 = current_row + 1

    # --- Header groupés ---
    ws.cell(row=header_row_1, column=2, value="Etudes")
    if not comparatif_mode:
        ws.cell(row=header_row_1, column=3, value=" Estimation")
        ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_1, end_column=6)
    # Col 1 (A1) incluse : fill obligatoire pour bloquer le débordement de B1 au scroll
    for c in range(1, 7):
        cell = ws.cell(row=header_row_1, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    # Badge LOT en première colonne : fond couleur du lot + texte "LOT XX" en blanc
    # Conforme aux modèles de référence qui identifient le lot dès la 1ère cellule.
    if _lot_num:
        _a1 = ws.cell(row=header_row_1, column=1, value=f"LOT {_lot_num}")
        _a1.font = Font(name="Tahoma", bold=True, size=10, color="FFFFFF")
        _a1.fill = PatternFill(start_color=_tab_color, end_color=_tab_color, fill_type="solid")
        _a1.alignment = Alignment(horizontal="center", vertical="center")

    company_start_col = 7
    for comp_idx, comp in enumerate(companies):
        start_col = company_start_col
        end_col = start_col + 4  # 5 colonnes : U. Qu. Px U. HT Px Tot HT Commentaire
        ws.cell(row=header_row_1, column=start_col, value=comp)
        ws.merge_cells(
            start_row=header_row_1, start_column=start_col, end_row=header_row_1, end_column=end_col
        )
        fill = FILL_COMPANY_COLORS[comp_idx % len(FILL_COMPANY_COLORS)]
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row_1, column=c)
            cell.font = FONT_HEADER_COMPANY
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        company_start_col = end_col + 1

    # --- Noms de colonnes ---
    base_headers = ["Code", "Désignation", "Qu.", "U", "Px U. HT", "Px Tot HT"]
    for i, header in enumerate(base_headers, 1):
        cell = ws.cell(row=header_row_2, column=i, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    col_offset = 7
    for comp_idx, _comp in enumerate(companies):
        fill = FILL_COMPANY_COLORS[comp_idx % len(FILL_COMPANY_COLORS)]
        for i, header in enumerate(["U.", "Qu.", "Px U. HT", "Px Tot HT", "Commentaire poste"]):
            cell = ws.cell(row=header_row_2, column=col_offset + i, value=header)
            cell.font = FONT_HEADER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        col_offset += 5

    max_col = 6 + len(companies) * 5

    # Index des alertes par code
    alert_by_code = {}
    for alert in alerts:
        code = alert.get("code", "")
        if code:
            alert_by_code.setdefault(code, []).append(alert)

    # --- PRE-COMPUTE : sections actives ---
    # Une section est "active" si elle contient au moins un article/sub_section avec
    # un montant non nul (template OU entreprise) ou une anomalie détectée.
    # Les sections 100% vides (ex : 01.3 … 01.10 sans données) sont ignorées à l'export.
    active_sections: set[str] = set()
    _cur_sec_scan: str | None = None
    for _, _row in merged_df.iterrows():
        _rt = _row["row_type"]
        if _rt == "section_header":
            _cur_sec_scan = str(_row.get("Code", "")).strip()
        elif _rt in ("article", "sub_section") and _cur_sec_scan:
            _code_scan = str(_row.get("Code", "")).strip()
            _has_data = False
            try:
                _has_data = float(_row.get("Px_Tot_HT", 0) or 0) != 0
            except (ValueError, TypeError):
                pass
            if not _has_data:
                for _comp in companies:
                    try:
                        if float(_row.get(f"{_comp}_Px_Tot_HT", 0) or 0) != 0:
                            _has_data = True
                            break
                    except (ValueError, TypeError):
                        pass
            if not _has_data and _code_scan and _code_scan in alert_by_code:
                _has_data = True
            if _has_data:
                active_sections.add(_cur_sec_scan)
    log.debug("Sections actives (%d) : %s", len(active_sections), sorted(active_sections))

    # --- ROWS 3+ : Données ---
    excel_row = 3

    # Tracking pour les formules dynamiques
    section_articles: dict[str, list[int]] = {}  # { '01.1': [row_idx, ...] }
    section_total_row: dict[str, int] = {}  # { '01.1': recap_row_idx }
    section_header_rows: dict[str, int] = {}  # { '01.1': section_header_row_idx }
    recap_summary_rows: list[int] = []  # [row_idx, ...]

    current_section_code: str | None = None
    _skip_section: bool = False  # True si la section courante est vide
    _recap_header_written: bool = False  # True dès que le bandeau RÉCAPITULATIF a été inséré

    # QW-3 : Variables initialisées ici (pas de 'in locals()' fragile)
    ht_row_idx: int | None = None
    tva_row_idx: int | None = None

    for _, row in merged_df.iterrows():
        row_type = row["row_type"]
        if row_type == "empty":
            continue

        code = str(row.get("Code", "")).strip()
        desig = str(row.get("Désignation", "")).strip()
        desig_lower = desig.lower()

        # OPT_DYN, 99 et SANS_CODE n'apparaissent pas dans le bandeau récapitulatif
        if row_type == "recap_summary" and code in ("OPT_DYN", "99", "SANS_CODE"):
            continue

        # Maj de la section courante + détection section vide
        if row_type == "section_header":
            # Si ce code a déjà été vu comme section_header, c'est la version
            # récapitulatif (apparition en fin de TCO) → le traiter comme recap_summary.
            if code in section_header_rows:
                row_type = "recap_summary"
            else:
                current_section_code = code
                _skip_section = code not in active_sections
                if not _skip_section:
                    if current_section_code not in section_articles:
                        section_articles[current_section_code] = []
                    section_header_rows[current_section_code] = excel_row

        # Ignorer les lignes d'une section sans données (articles/recap internes)
        # Les recap_summary sont toujours affichés — même à 0 — pour un récapitulatif complet.
        if _skip_section and row_type in ("section_header", "sub_section", "article", "recap"):
            continue

        # ── BANDEAU "RÉCAPITULATIF" avant la première ligne recap_summary ──
        if row_type == "recap_summary" and not _recap_header_written:
            _recap_header_written = True
            # Ligne de séparation fine (blanche)
            for c in range(1, max_col + 1):
                ws.cell(row=excel_row, column=c).fill = FILL_WHITE
            ws.row_dimensions[excel_row].height = 6
            excel_row += 1
            # Ligne séparatrice colorée (trait bleu) — visible entre le détail et le récap
            for c in range(1, max_col + 1):
                cell = ws.cell(row=excel_row, column=c)
                cell.fill = FILL_RECAP_SEPARATOR
                cell.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))
            ws.row_dimensions[excel_row].height = 4
            excel_row += 1
            # Bandeau titre principal — fond marine, texte blanc large et centré
            for c in range(1, max_col + 1):
                cell = ws.cell(row=excel_row, column=c)
                cell.fill = FILL_RECAP_HEADER
                cell.font = FONT_RECAP_HEADER_LARGE
                cell.border = MEDIUM_BORDER
            title_cell = ws.cell(row=excel_row, column=1, value="  📋  RÉCAPITULATIF")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[excel_row].height = 28
            excel_row += 1
        # ────────────────────────────────────────────────────────────────────────

        ws.cell(row=excel_row, column=1, value=code)
        ws.cell(row=excel_row, column=2, value=row["Désignation"])
        # section_header = titre de section, pas de données chiffrées
        if row_type != "section_header":
            ws.cell(row=excel_row, column=3, value=_clean_val(row.get("Qu.")))
            ws.cell(row=excel_row, column=4, value=row.get("U"))
            ws.cell(row=excel_row, column=5, value=_clean_val(row.get("Px_U_HT")))

        # --- Colonne F : TCO / Estimation ---
        # FIX CAS 2/3/4 : sub_section ET article contribuent au total de leur section.
        # Les sub_sections (Entete _Niv1/_Niv2) ont des prix propres (ex: 06.5.3.2)
        # et doivent apparaître dans le SUM recap — elles étaient silencieusement omises.
        if row_type in ("article", "sub_section") and current_section_code:
            section_articles[current_section_code].append(excel_row)

        if row_type == "article":
            qu_val = _clean_val(row.get("Qu."))
            px_val = _clean_val(row.get("Px_U_HT"))
            if isinstance(qu_val, (int, float)) and isinstance(px_val, (int, float)):
                ws.cell(row=excel_row, column=6, value=f"=C{excel_row}*E{excel_row}")
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))

        elif row_type == "recap":
            # Total de la section (ligne grise)
            rows = section_articles.get(current_section_code, [])
            if rows:
                ws.cell(row=excel_row, column=6, value=_rows_to_sum_formula("F", rows))
            else:
                # Section forfaitaire (code ≤2 segments avec prix direct, sans articles enfants).
                # La valeur a été propagée par compute_section_totals Passe 2 vers cette ligne recap.
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))
            section_total_row[current_section_code] = excel_row

        elif row_type == "recap_summary":
            # Ligne dans le tableau final récapitulatif
            recap_summary_rows.append(excel_row)
            # Cascade : 1) recap interne, 2) SUM articles, 3) valeur statique
            target_row = section_total_row.get(code)
            art_rows = section_articles.get(code, [])
            if target_row:
                ws.cell(row=excel_row, column=6, value=f"=F{target_row}")
            elif art_rows:
                ws.cell(row=excel_row, column=6, value=_rows_to_sum_formula("F", art_rows))
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))

        elif _RE_MONTANT_HT.search(desig_lower):
            # Grand Total HT — somme des lignes récapitulatif
            # Priorité : recap_summary_rows, sinon section_total_row (fallback PDF)
            _sum_rows = recap_summary_rows or list(section_total_row.values())
            if _sum_rows:
                ws.cell(row=excel_row, column=6, value=_rows_to_sum_formula("F", _sum_rows))
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))
            ht_row_idx = excel_row
        elif _RE_TVA_ONLY.search(desig_lower) and not _RE_HT_ONLY.search(desig_lower):
            # TVA = HT * taux
            if ht_row_idx is not None:
                ws.cell(row=excel_row, column=6, value=f"=F{ht_row_idx}*{tva_rate}")
                tva_row_idx = excel_row
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))
        elif _RE_MONTANT_TTC.search(desig_lower):
            # TTC = HT + TVA
            if ht_row_idx is not None and tva_row_idx is not None:
                ws.cell(row=excel_row, column=6, value=f"=F{ht_row_idx}+F{tva_row_idx}")
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))
        elif row_type == "sub_section":
            # PARTIE 3 : formule =C*E si Qu. et Px_U_HT présents, sinon montant merger
            try:
                qu_val = float(row.get("Qu.", 0) or 0)
                px_u_val = float(row.get("Px_U_HT", 0) or 0)
            except (ValueError, TypeError):
                qu_val = px_u_val = 0.0
            if qu_val != 0 and px_u_val != 0:
                ws.cell(row=excel_row, column=6, value=f"=C{excel_row}*E{excel_row}")
            else:
                ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))
        elif row_type == "section_header":
            # Titre de section : pas de montant affiché — le total est porté
            # par la ligne recap en fin de section (évite le doublon visuel).
            ws.cell(row=excel_row, column=6, value=None)
        else:
            ws.cell(row=excel_row, column=6, value=_clean_val(row.get("Px_Tot_HT")))

        # --- Colonnes Entreprises ---
        col_offset = 7
        for comp in companies:
            if row_type != "section_header":
                ws.cell(row=excel_row, column=col_offset, value=row.get(f"{comp}_U."))
                ws.cell(row=excel_row, column=col_offset + 1, value=_clean_val(row.get(f"{comp}_Qu.")))
                ws.cell(row=excel_row, column=col_offset + 2, value=_clean_val(row.get(f"{comp}_Px_U_HT")))

            qu_col = get_column_letter(col_offset + 1)
            px_col = get_column_letter(col_offset + 2)
            tot_col = get_column_letter(col_offset + 3)

            if row_type == "article":
                comp_qu = _clean_val(row.get(f"{comp}_Qu."))
                comp_px = _clean_val(row.get(f"{comp}_Px_U_HT"))
                if isinstance(comp_qu, (int, float)) and isinstance(comp_px, (int, float)):
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=f"={qu_col}{excel_row}*{px_col}{excel_row}",
                    )
                else:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=_clean_val(row.get(f"{comp}_Px_Tot_HT")),
                    )
            elif row_type == "recap":
                rows = section_articles.get(current_section_code, [])
                if rows:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=_rows_to_sum_formula(tot_col, rows),
                    )
                else:
                    # Section forfaitaire : valeur propagée par Passe 2 sur la ligne recap.
                    ws.cell(row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT")))
            elif row_type == "recap_summary":
                target_row = section_total_row.get(code)
                art_rows = section_articles.get(code, [])
                if target_row:
                    ws.cell(row=excel_row, column=col_offset + 3, value=f"={tot_col}{target_row}")
                elif art_rows:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=_rows_to_sum_formula(tot_col, art_rows),
                    )
                else:
                    ws.cell(
                        row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT"))
                    )
            elif _RE_MONTANT_HT.search(desig_lower):
                _sum_rows = recap_summary_rows or list(section_total_row.values())
                if _sum_rows:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=_rows_to_sum_formula(tot_col, _sum_rows),
                    )
                else:
                    ws.cell(
                        row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT"))
                    )
            elif _RE_TVA_ONLY.search(desig_lower) and not _RE_HT_ONLY.search(desig_lower):
                if ht_row_idx is not None:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=f"={tot_col}{ht_row_idx}*{tva_rate}",
                    )
                else:
                    ws.cell(
                        row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT"))
                    )
            elif _RE_MONTANT_TTC.search(desig_lower):
                if ht_row_idx is not None and tva_row_idx is not None:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=f"={tot_col}{ht_row_idx}+{tot_col}{tva_row_idx}",
                    )
                else:
                    ws.cell(
                        row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT"))
                    )
            elif row_type == "sub_section":
                comp_qu = _clean_val(row.get(f"{comp}_Qu."))
                comp_px = _clean_val(row.get(f"{comp}_Px_U_HT"))
                if isinstance(comp_qu, (int, float)) and isinstance(comp_px, (int, float)) and comp_qu != 0 and comp_px != 0:
                    ws.cell(
                        row=excel_row,
                        column=col_offset + 3,
                        value=f"={qu_col}{excel_row}*{px_col}{excel_row}",
                    )
                else:
                    ws.cell(
                        row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT"))
                    )
            elif row_type == "section_header":
                ws.cell(row=excel_row, column=col_offset + 3, value=None)
            else:
                ws.cell(row=excel_row, column=col_offset + 3, value=_clean_val(row.get(f"{comp}_Px_Tot_HT")))

            if row_type != "section_header":
                ws.cell(row=excel_row, column=col_offset + 4, value=row.get(f"{comp}_Commentaire"))
            col_offset += 5

        # --- Format numérique (appliqué à toutes les lignes) ---
        ws.cell(row=excel_row, column=3).number_format = QTY_FORMAT  # Qu. TCO
        ws.cell(row=excel_row, column=5).number_format = MONEY_FORMAT  # Px U HT TCO
        ws.cell(row=excel_row, column=6).number_format = MONEY_FORMAT  # Px Tot HT TCO
        _ncol = 7
        for _ in companies:
            # col +0 = U. (texte) — pas de format numérique
            ws.cell(row=excel_row, column=_ncol + 1).number_format = QTY_FORMAT  # Qu.
            ws.cell(row=excel_row, column=_ncol + 2).number_format = MONEY_FORMAT  # Px U HT
            ws.cell(row=excel_row, column=_ncol + 3).number_format = MONEY_FORMAT  # Px Tot HT
            _ncol += 5

        # --- Style ---
        style_type = row_type
        if row_type == "section_header":
            if code == "OPT_DYN":
                style_type = "class_opt"
            elif code == "99":
                style_type = "class_99"
            elif code == "SANS_CODE":
                style_type = "class_sans_code"
        elif row_type == "recap":
            if code in ("OPT_DYN", "99", "SANS_CODE"):
                style_type = "recap"  # Style recap standard pour les sections dynamiques

        font, fill = _get_row_style(style_type)

        # PARTIE 2 : sub_section sans prix → titre principal (BATIMENT F, etc.)
        if row_type == "sub_section":
            try:
                px_val = float(row.get("Px_Tot_HT", 0) or 0)
            except (ValueError, TypeError):
                px_val = 0.0
            if px_val == 0:
                font, fill = FONT_MAIN_TITLE, FILL_MAIN_TITLE

        # Hiérarchie visuelle : surcharge pour les lignes totaux généraux
        if row_type == "total_line":
            if _RE_MONTANT_HT.search(desig_lower):
                font, fill = FONT_GRAND_TOTAL, FILL_MONTANT_HT
            elif _RE_TVA_ONLY.search(desig_lower) and not _RE_HT_ONLY.search(desig_lower):
                font, fill = FONT_GRAND_TOTAL, FILL_TVA
            elif _RE_MONTANT_TTC.search(desig_lower):
                font, fill = FONT_GRAND_TOTAL, FILL_MONTANT_TTC

        # Sous-totaux intermédiaires (recap de section imbriquée) : 8pt comme la référence.
        # Un recap de section de niveau ≥ 2 (ex: "01.2.1") a 2 points dans son code
        # et reçoit une police plus petite que les totaux de section principale (11pt).
        if row_type == "recap" and code.count(".") >= 2:
            font = FONT_RECAP_SUB

        # Articles alternatifs : code suffixé par une ou plusieurs lettres après un chiffre
        # (ex: "01.1A", "10.2.2.1.4.1A", "11.2.1.1.2AA") → orange comme les modèles.
        if row_type in ("article", "sub_section", "section_header"):
            if _RE_ALT_ARTICLE.search(code):
                font = FONT_ALTERNATIVE

        # --- OPTIONS STYLING ---
        if row.get("is_option") is True:
            # On applique l'italique pour toutes les lignes marquées option
            font = _copy(font)
            font.italic = True

        # --- LIGNE AJOUTÉE PAR L'ENTREPRISE ---
        if row.get("is_added") is True:
            fill = FILL_ADDED

        # --- ZEBRA STRIPING DISCRET ---
        # S'applique uniquement aux articles pour aider à la lecture horizontale
        if row_type == "article" and (excel_row % 2 == 0) and not fill:
            fill = FILL_ZEBRA_EVEN

        # --- OUTLINE LEVELS (Groupement pliable) ---
        # On définit le niveau de profondeur basé sur le nombre de points dans le code
        # Ex: "01" = 0, "01.1" = 1, "01.1.1" = 2. Plafond à 7 (limite Excel).
        if row_type in ("article", "sub_section"):
            _lvl = min(7, code.count("."))
            if _lvl > 0:
                ws.row_dimensions[excel_row].outline_level = _lvl

        # Récapitulatif : bordure épaisse pour encadrer chaque ligne
        _border = MEDIUM_BORDER if row_type in ("recap_summary", "recap") else THIN_BORDER

        # --- DÉTECTION DU MOINS DISANT (Meilleur Prix) ---
        min_pu = None
        if row_type == "article":
            all_pus = []
            for comp in companies:
                try:
                    p = float(row.get(f"{comp}_Px_U_HT", 0) or 0)
                    if p > 0:
                        all_pus.append(p)
                except (ValueError, TypeError):
                    pass
            if all_pus:
                min_pu = min(all_pus)

        for c in range(1, max_col + 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = font
            cell.border = _border
            cell.fill = fill if fill else FILL_WHITE
            cell.alignment = Alignment(vertical="center")

        # Bordure épaisse pour séparer l'Estimation TCO des Entreprises
        ws.cell(row=excel_row, column=6).border = THICK_RIGHT_BORDER

        # --- BORDURES ÉPAISSES ENTRE ENTREPRISES ET MOINS DISANT ---
        # On repasse sur les colonnes entreprises pour les finitions
        _c_off = 7
        for comp in companies:
            # Bordure droite plus épaisse à la fin de chaque bloc entreprise (+4 = Commentaire)
            ws.cell(row=excel_row, column=_c_off + 4).border = THICK_RIGHT_BORDER

            # Highlight Moins Disant
            if min_pu is not None:
                try:
                    raw_val = row.get(f"{comp}_Px_U_HT")
                    curr_pu = float(raw_val or 0)

                    if abs(curr_pu - min_pu) < 0.001:
                        # On met le prix unitaire en gras / couleur pour le distinguer
                        best_font = _copy(cell.font)
                        best_font.bold = True
                        best_font.color = "38761D" # Vert foncé pro
                        ws.cell(row=excel_row, column=_c_off + 2).font = best_font

                    # --- NON CHIFFRÉ (NC) ---
                    # On n'affiche NC que si la valeur est VRAIMENT absente (None ou NaN)
                    # et NON SI elle vaut 0 (qui peut être "Compris").
                    # On évite de l'écrire si c'est une ligne recap/total.
                    if pd.isna(raw_val) and row_type == "article":
                        # Colonne Prix Unitaire
                        nc_cell = ws.cell(row=excel_row, column=_c_off + 2)
                        nc_cell.value = "NC"
                        nc_cell.font = Font(name="Tahoma", size=7, color="999999", italic=True)
                        nc_cell.alignment = Alignment(horizontal="center", vertical="center")

                        # Colonne Prix Total : on efface la formule Qu*PU qui renverrait #VALEUR!
                        ws.cell(row=excel_row, column=_c_off + 3, value=None)
                except (ValueError, TypeError):
                    pass

            _c_off += 5

        # recap_summary : teinter les colonnes de chaque entreprise avec leur couleur.
        # Les colonnes A–F gardent FILL_RECAP_SUMMARY ; chaque groupe de 5 colonnes
        # entreprise reçoit la teinte claire correspondant à la couleur de son en-tête.
        if row_type == "recap_summary":
            _tint_off = 7
            for _tint_idx in range(len(companies)):
                _tint = FILL_COMPANY_TINTS[_tint_idx % len(FILL_COMPANY_TINTS)]
                for _tc in range(_tint_off, _tint_off + 5):
                    ws.cell(row=excel_row, column=_tc).fill = _tint
                _tint_off += 5

        # Col A (Code) : left + center — même ancrage vertical que col B.
        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

        # Col B (Désignation) : wrap_text pour les textes longs + vertical center
        # pour rester aligné avec col A quelle que soit la hauteur de ligne.
        _indent = 0
        if row_type == "article":
            _indent = 2
        elif row_type == "sub_section" and font is not FONT_MAIN_TITLE:
            _indent = 1
        ws.cell(row=excel_row, column=2).alignment = Alignment(
            horizontal="left", indent=_indent, wrap_text=True, vertical="center"
        )

        # Cols Commentaire : wrap_text pour éviter que les longs débordent sur les colonnes adjacentes
        _c_off = 7
        for _ in companies:
            ws.cell(row=excel_row, column=_c_off + 4).alignment = Alignment(
                horizontal="left", wrap_text=True, vertical="center"
            )
            _c_off += 5

        if code and code in alert_by_code and row_type in ("article", "sub_section"):
            # Collecter la sévérité max + les entreprises concernées
            max_severity = "info"
            companies_with_errors: set[str] = set()
            companies_with_warnings: set[str] = set()
            has_global_error = False

            for alert in alert_by_code[code]:
                sev = alert["type"]
                comp = alert.get("company", "")
                if sev == "error":
                    max_severity = "error"
                    if comp:
                        companies_with_errors.add(comp)
                    else:
                        has_global_error = True
                elif sev == "warning" and max_severity != "error":
                    max_severity = "warning"
                    if comp:
                        companies_with_warnings.add(comp)

            if max_severity == "error":
                # Collect error messages for Commentaire
                err_msgs = [a["message"] for a in alert_by_code[code] if a["type"] == "error"]
                err_text = " | ".join(err_msgs)
                if has_global_error or not companies_with_errors:
                    # Erreur non liée à une entreprise → ligne entière
                    for c in range(1, max_col + 1):
                        ws.cell(row=excel_row, column=c).fill = FILL_ERROR
                    # Écrire la raison dans la colonne Commentaire de chaque entreprise
                    _c_off = 7
                    for _ in companies:
                        _com_cell = ws.cell(row=excel_row, column=_c_off + 4)
                        if not _com_cell.value:
                            _com_cell.value = f"⚠ {err_text}"
                        _c_off += 5
                else:
                    # Erreur ciblée → seulement les 5 colonnes de l'entreprise fautive
                    for comp in companies_with_errors:
                        if comp in companies:
                            _ci = companies.index(comp)
                            _sc = 7 + _ci * 5
                            for c in range(_sc, _sc + 5):
                                ws.cell(row=excel_row, column=c).fill = FILL_ERROR
                            # Message dans Commentaire de cette entreprise
                            _com_cell = ws.cell(row=excel_row, column=_sc + 4)
                            if not _com_cell.value:
                                comp_msgs = [
                                    a["message"]
                                    for a in alert_by_code[code]
                                    if a["type"] == "error" and a.get("company", "") == comp
                                ]
                                if comp_msgs:
                                    _com_cell.value = f"⚠ {' | '.join(comp_msgs)}"

            elif max_severity == "warning":
                # Collect warning messages
                warn_msgs = [a["message"] for a in alert_by_code[code] if a["type"] == "warning"]
                warn_text = " | ".join(warn_msgs)
                if companies_with_warnings:
                    # Warning ciblé → 5 colonnes de l'entreprise concernée
                    for comp in companies_with_warnings:
                        if comp in companies:
                            _ci = companies.index(comp)
                            _sc = 7 + _ci * 5
                            for c in range(_sc, _sc + 5):
                                ws.cell(row=excel_row, column=c).fill = FILL_WARNING
                            # Message dans Commentaire de l'entreprise
                            _com_cell = ws.cell(row=excel_row, column=_sc + 4)
                            if not _com_cell.value:
                                comp_msgs = [
                                    a["message"]
                                    for a in alert_by_code[code]
                                    if a["type"] == "warning" and a.get("company", "") == comp
                                ]
                                if comp_msgs:
                                    _com_cell.value = f"⚠ {' | '.join(comp_msgs)}"
                else:
                    # Fallback global : colonne Commentaire de toutes les entreprises
                    _c_off = 7
                    for _ in companies:
                        ws.cell(row=excel_row, column=_c_off + 4).fill = FILL_WARNING
                        _com_cell = ws.cell(row=excel_row, column=_c_off + 4)
                        if not _com_cell.value:
                            _com_cell.value = f"⚠ {warn_text}"
                        _c_off += 5

        # Hauteur de ligne dynamique : s'adapte aux désignations longues qui wrappent.
        # Col B = 56.75 chars → ~55 chars utiles par ligne à 9pt.
        # Chaque ligne de texte ≈ 16 pt ; plancher 28.5 pt (1 ligne), plafond 80 pt.
        # Les lignes recap_summary ont une hauteur fixe légèrement plus grande pour
        # les différencier visuellement du reste du tableau.
        if row_type in ("recap_summary", "recap"):
            ws.row_dimensions[excel_row].height = 22.0
        else:
            _n_lines_desig = max(1, -(-len(desig) // 55)) if desig else 1

            # Examiner la longueur des commentaires pour adapter la hauteur
            _max_com_len = 0
            _c_off = 7
            for _ in companies:
                _com_val = ws.cell(row=excel_row, column=_c_off + 4).value
                if _com_val:
                    _max_com_len = max(_max_com_len, len(str(_com_val)))
                _c_off += 5

            # Une ligne de commentaire fait environ 30 caractères utiles pour col.width=30
            _n_lines_com = max(1, -(-_max_com_len // 30)) if _max_com_len > 0 else 1
            _n_lines = max(_n_lines_desig, _n_lines_com)

            # Plafond augmenté à 150 pt pour bien lire les gros paquets d'erreurs
            ws.row_dimensions[excel_row].height = max(28.5, min(_n_lines * 16.0, 150.0))

        excel_row += 1

    # PARTIE 3 : Injection différée des formules pour section_header
    # Les plages d'articles/sub_sections ne sont connues qu'après le parcours complet.
    for sh_code, sh_row in section_header_rows.items():
        recap_row = section_total_row.get(sh_code)
        art_rows = section_articles.get(sh_code, [])

        # Col F : référencer le recap (=F{recap_row}) ou sommer les enfants
        if recap_row:
            ws.cell(row=sh_row, column=6, value=f"=F{recap_row}")
        elif art_rows:
            ws.cell(row=sh_row, column=6, value=_rows_to_sum_formula("F", art_rows))

        # Colonnes entreprises
        c_off = 7
        for _ in companies:
            tc = get_column_letter(c_off + 3)  # Px Tot HT est maintenant col +3
            if recap_row:
                ws.cell(row=sh_row, column=c_off + 3, value=f"={tc}{recap_row}")
            elif art_rows:
                ws.cell(row=sh_row, column=c_off + 3, value=_rows_to_sum_formula(tc, art_rows))
            c_off += 5

    # Largeurs de colonnes — valeurs exactes de la référence
    ws.column_dimensions["A"].width = 9.5
    ws.column_dimensions["B"].width = 56.75
    ws.column_dimensions["C"].width = 9.5
    ws.column_dimensions["D"].width = 7.125
    ws.column_dimensions["E"].width = 14.125
    ws.column_dimensions["F"].width = 16.5
    for _ci in range(len(companies)):
        _cb = 7 + _ci * 5
        ws.column_dimensions[get_column_letter(_cb)].width = 6.0  # U.
        ws.column_dimensions[get_column_letter(_cb + 1)].width = 9.5  # Qu.
        ws.column_dimensions[get_column_letter(_cb + 2)].width = 14.125  # Px U HT
        ws.column_dimensions[get_column_letter(_cb + 3)].width = 16.5  # Px Tot HT
        ws.column_dimensions[get_column_letter(_cb + 4)].width = 30.0  # Commentaire (élargi pour wrap)

    # Hauteurs en-têtes : 14.25 pt (conforme référence)
    ws.row_dimensions[header_row_1].height = 14.25
    ws.row_dimensions[header_row_2].height = 14.25

    # Mode comparatif : masquer les colonnes Estimation (C=Qu. D=U E=Px U HT F=Px Tot HT)
    if comparatif_mode:
        for _col_letter in ("C", "D", "E", "F"):
            ws.column_dimensions[_col_letter].hidden = True

    # Freeze panes : lignes d'en-tête + colonnes A (Code) et B (Désignation) figées
    fix_freeze_panes(
        ws, header_rows=header_row_2, frozen_cols=2
    )  # C{header_row_2 + 1} : lignes 1-header_row_2 + cols A-B figées
    fix_merged_cells_crossing_freeze(
        ws, header_rows=header_row_2, frozen_cols=2
    )  # retire fusions qui traversent la frontière de freeze
    prevent_text_overflow(
        ws, min_row=header_row_2 + 1, max_col=max_col
    )  # fill blanc sur cellules vides

    log.info("Workbook prêt. Output_path=%s", output_path)

    if output_path:
        wb.save(output_path)
        wb.close()
        return output_path
    else:
        buffer = io.BytesIO()
        wb.save(buffer)
        wb.close()
        buffer.seek(0)
        return buffer
