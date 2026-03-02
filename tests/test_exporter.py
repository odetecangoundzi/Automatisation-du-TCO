"""
test_exporter.py — Tests de core/exporter.py.

Vérifie le comportement observable de export_tco via le BytesIO/fichier retourné.
"""

from __future__ import annotations

import io
import os
from decimal import Decimal

import openpyxl
import pandas as pd
import pytest

from core.exporter import export_tco

# ---------------------------------------------------------------------------
# Fixture locale
# ---------------------------------------------------------------------------


@pytest.fixture
def merged_df_one_company(minimal_tco_df):
    """TCO avec colonnes entreprise pré-remplies (simule le résultat de merge)."""
    df = minimal_tco_df.copy()
    df["ACME_Qu."] = [None, Decimal("10"), None]
    df["ACME_Px_U_HT"] = [None, Decimal("25"), None]
    df["ACME_Px_Tot_HT"] = [Decimal("250"), Decimal("250"), Decimal("250")]
    df["ACME_Commentaire"] = [None, "", None]
    return df


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestExportTCO:
    def test_export_returns_bytesio_when_no_output_path(self, merged_df_one_company):
        """Sans output_path → retourne un BytesIO."""
        result = export_tco(merged_df_one_company, meta={})
        assert isinstance(result, io.BytesIO)
        # Le BytesIO doit contenir des données
        result.seek(0)
        content = result.read()
        assert len(content) > 100

    def test_export_creates_file_when_output_path_given(self, merged_df_one_company, tmp_path):
        """Avec output_path → fichier créé sur disque."""
        out = str(tmp_path / "out.xlsx")
        result = export_tco(merged_df_one_company, meta={}, output_path=out)
        assert isinstance(result, str)
        assert os.path.exists(result)
        assert os.path.getsize(result) > 100

    def test_export_sheet_has_data(self, merged_df_one_company):
        """La feuille exportée contient des cellules non vides."""
        buf = export_tco(merged_df_one_company, meta={})
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        # Au moins quelques cellules doivent avoir des valeurs
        non_empty = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None
        ]
        assert len(non_empty) > 0

    def test_export_company_name_in_header(self, merged_df_one_company):
        """Le nom d'entreprise 'ACME' apparaît dans les headers (row 1 ou 2)."""
        buf = export_tco(merged_df_one_company, meta={})
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        found = False
        for r in range(1, 4):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and "ACME" in str(val):
                    found = True
                    break
        assert found, "Le nom 'ACME' devrait apparaître dans les premières lignes"

    def test_export_readable_by_pandas(self, merged_df_one_company):
        """Le BytesIO est lisible par pandas sans exception."""
        buf = export_tco(merged_df_one_company, meta={})
        buf.seek(0)
        df_back = pd.read_excel(buf, engine="openpyxl")
        assert isinstance(df_back, pd.DataFrame)

    def test_export_with_alerts_no_crash(self, merged_df_one_company):
        """L'export avec des alertes ne lève pas d'exception."""
        alerts = [
            {"type": "error", "color": "red", "row": 3, "code": "1.1", "message": "test error"}
        ]
        result = export_tco(merged_df_one_company, meta={}, alerts=alerts)
        assert result is not None

    def test_export_empty_df_no_crash(self):
        """DataFrame vide → pas d'exception (résilience)."""
        try:
            result = export_tco(pd.DataFrame(), meta={})
            # Si ça retourne quelque chose, c'est OK
            assert result is not None
        except Exception as e:
            pytest.fail(f"export_tco a levé une exception inattendue : {e}")

    def test_export_with_meta_project_info(self, merged_df_one_company):
        """Export avec meta project_info ne lève pas d'exception."""
        meta = {
            "project_info": {"projet": "TEST", "lot": "01"},
            "header_row": 5,
            "sheet_name": "TCO",
            "filepath": "test.xlsx",
        }
        result = export_tco(merged_df_one_company, meta=meta, tva_rate=0.20)
        assert isinstance(result, io.BytesIO)
