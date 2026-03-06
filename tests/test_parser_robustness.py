import openpyxl
import pytest

from core.merger import merge_company_into_tco
from core.parser_dpgf import parse_dpgf
from core.parser_tco import parse_tco

# -----------------
# Fixtures & Utils
# -----------------


@pytest.fixture
def empty_excel(tmp_path):
    fpath = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(fpath)
    return str(fpath)


@pytest.fixture
def no_header_excel(tmp_path):
    fpath = tmp_path / "no_header.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Coucou", "Ceci", "N'est", "Pas", "Un", "DPGF"])
    ws.append(["123", "456", "789"])
    wb.save(fpath)
    return str(fpath)


@pytest.fixture
def valid_tco(tmp_path):
    fpath = tmp_path / "valid_tco.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["PROJET TEST"])
    ws.append(["Code", "Désignation", "Qu.", "U", "Px U", "Px Tot", "Entete"])
    ws.append(["01", "Section 1", "", "", "", "", "Bd_01_Bord"])
    ws.append(["01.1.1", "Article 1", "10", "u", "10", "100", "Ouv_01_Art"])
    wb.save(fpath)
    return str(fpath)


@pytest.fixture
def valid_dpgf(tmp_path):
    fpath = tmp_path / "valid_dpgf.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Code", "Désignation", "Qu.", "U", "Px U", "Px Tot"])
    ws.append(["01.1.1", "Article 1 - DPGF", "10", "u", "15", "150"])
    wb.save(fpath)
    return str(fpath)


# -------------
# Tests TCO
# -------------


def test_parse_tco_empty(empty_excel):
    """S'assure que parser un TCO vide ne crash pas."""
    df, meta = parse_tco(empty_excel)
    assert df.empty, "Le DataFrame devrait être vide"
    assert "error" in meta, "Les métadonnées devraient contenir une clé d'erreur"
    assert "Impossible de trouver la ligne d'en-tête" in meta["error"]


def test_parse_tco_no_header(no_header_excel):
    """S'assure que parser un TCO sans header ne crash pas."""
    df, meta = parse_tco(no_header_excel)
    assert df.empty, "Le DataFrame devrait être vide"
    assert "error" in meta, "Les métadonnées devraient contenir une clé d'erreur"


def test_parse_tco_valid(valid_tco):
    df, meta = parse_tco(valid_tco)
    assert not df.empty
    assert len(df) == 2
    assert "01.1.1" in df["Code"].values


# -------------
# Tests DPGF
# -------------


def test_parse_dpgf_empty(empty_excel):
    """S'assure que parser un DPGF vide ne crash pas."""
    df, alerts = parse_dpgf(empty_excel)
    assert df.empty, "Le DataFrame devrait être vide"
    assert len(alerts) > 0, "Devrait retourner des alertes"
    assert "Impossible de trouver la ligne" in alerts[0]["message"]


def test_parse_dpgf_no_header(no_header_excel):
    """S'assure que parser un DPGF sans header ne crash pas."""
    df, alerts = parse_dpgf(no_header_excel)
    assert df.empty, "Le DataFrame devrait être vide"
    assert len(alerts) > 0, "Devrait retourner des alertes"


def test_parse_dpgf_valid(valid_dpgf):
    df, alerts = parse_dpgf(valid_dpgf)
    assert not df.empty
    assert len(df) == 1
    assert "01.1.1" in df["Code"].values


# -------------
# Tests Merge
# -------------


def test_merge_empty_dpgf(valid_tco, empty_excel):
    """Si le TCO est valide mais que le DPGF est corrompu/vide, la fusion ne doit pas crasher."""
    tco_df, _ = parse_tco(valid_tco)
    dpgf_df, _ = parse_dpgf(empty_excel)

    merged_df, alerts = merge_company_into_tco(tco_df, dpgf_df, "ENTREPRISE_X")

    assert not merged_df.empty, (
        "Le DataFrame fusionné devrait contenir le TCO même si le DPGF est vide"
    )
    assert "ENTREPRISE_X_Px_Tot_HT" in merged_df.columns
    assert len(alerts) == 1, "Devrait retourner une alerte indiquant que le DPGF est vide"
    assert "vide" in alerts[0]["message"].lower()


def test_merge_valid(valid_tco, valid_dpgf):
    tco_df, _ = parse_tco(valid_tco)
    dpgf_df, _ = parse_dpgf(valid_dpgf)

    merged_df, alerts = merge_company_into_tco(tco_df, dpgf_df, "TEST_COMP")

    # L'article 01.1.1 doit avoir reçu le prix 150
    article_row = merged_df[merged_df["Code"] == "01.1.1"].iloc[0]
    assert article_row["TEST_COMP_Px_Tot_HT"] == 150
    assert article_row["TEST_COMP_Px_U_HT"] == 15

    # La section_header 01 est vidée (doublon éliminé), le recap porte le total
    section_row = merged_df[merged_df["Code"] == "01"].iloc[0]
    assert section_row["TEST_COMP_Px_Tot_HT"] is None
